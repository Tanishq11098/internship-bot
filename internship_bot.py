import smtplib, os, time, random
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup

YOUR_EMAIL    = os.environ["EMAIL"]
YOUR_PASSWORD = os.environ["PASSWORD"]
SEND_TO_EMAIL = os.environ["EMAIL"]

HEADERS_LIST = [
    {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"},
    {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/119.0.0.0 Safari/537.36"},
    {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/118.0.0.0 Safari/537.36"},
]

DOMAIN_COLORS = {
    "Investment Banking": "C00000",
    "Equity Research":    "E26B0A",
    "Wealth Management":  "375623",
    "Private Equity":     "7030A0",
    "Venture Capital":    "7030A0",
    "Hedge Fund":         "7030A0",
    "Credit / Debt":      "1F4E79",
    "Treasury":           "1F4E79",
    "Compliance / Risk":  "833C00",
    "Audit / Assurance":  "833C00",
    "Valuation":          "E26B0A",
    "FP&A":               "375623",
    "Fintech / Payments": "0070C0",
    "Finance Operations": "404040",
    "Accounting / Tax":   "404040",
    "Asset Management":   "375623",
    "Portfolio Management":"375623",
    "KPO Finance":        "1F4E79",
}

def get_headers():
    return random.choice(HEADERS_LIST)

def detect_domain(title, company=""):
    text = (title + " " + company).lower()
    if any(x in text for x in ["investment bank", "ib ", "m&a", "capital market", "corporate finance"]): return "Investment Banking"
    if any(x in text for x in ["equity research", "equity analyst", "stock analyst", "market research", "research analyst"]): return "Equity Research"
    if any(x in text for x in ["wealth", "private banking", "relationship manager", "hni"]): return "Wealth Management"
    if any(x in text for x in ["private equity", " pe ", "buyout", "growth equity"]): return "Private Equity"
    if any(x in text for x in ["venture", " vc ", "startup invest"]): return "Venture Capital"
    if any(x in text for x in ["hedge", "quant", "algo", "derivatives"]): return "Hedge Fund"
    if any(x in text for x in ["credit", "debt", "lending", "fixed income", "bond", "nbfc"]): return "Credit / Debt"
    if any(x in text for x in ["treasury", "forex", " fx ", "currency", "liquidity"]): return "Treasury"
    if any(x in text for x in ["compliance", " risk", "regulatory", "kyc", "aml"]): return "Compliance / Risk"
    if any(x in text for x in ["audit", "assurance", "internal audit"]): return "Audit / Assurance"
    if any(x in text for x in ["valuat", "dcf", "modell", "financial model"]): return "Valuation"
    if any(x in text for x in ["fp&a", "financial planning", "budgeting", "forecasting"]): return "FP&A"
    if any(x in text for x in ["fintech", "payment", "wallet", "insurtech", "neobank"]): return "Fintech / Payments"
    if any(x in text for x in ["kpo", "research process"]): return "KPO Finance"
    if any(x in text for x in ["portfolio", "asset manag", "fund manag", "aum", "mutual fund"]): return "Asset Management"
    if any(x in text for x in ["account", " tax", "gst", "tally", "bookkeep"]): return "Accounting / Tax"
    return "Finance Operations"

# ─────────────────────────────────────────────
# SOURCE 1: INTERNSHALA
# ─────────────────────────────────────────────
def scrape_internshala():
    results = []
    urls = [
        ("https://internshala.com/internships/finance-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/finance-internship-in-gurgaon/", "Gurugram"),
        ("https://internshala.com/internships/finance-internship-in-noida/", "Noida"),
        ("https://internshala.com/internships/investment-banking-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/equity-research-internship/", "Delhi NCR"),
        ("https://internshala.com/internships/audit-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/mba-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/accounting-internship-in-delhi/", "Delhi"),
    ]
    for url, loc in urls:
        try:
            time.sleep(random.uniform(1, 2))
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.select(".internship_meta")[:20]
            for card in cards:
                title    = card.select_one(".profile")
                company  = card.select_one(".company_name")
                location = card.select_one(".location_link")
                stipend  = card.select_one(".stipend")
                duration = card.select_one(".item_body")
                deadline = card.select_one(".apply-by")
                link_tag = card.find_parent("a")
                if title:
                    t = title.get_text(strip=True)
                    c = company.get_text(strip=True) if company else "N/A"
                    results.append({
                        "title"    : t,
                        "company"  : c,
                        "firm_type": "Startup / Corporate",
                        "domain"   : detect_domain(t, c),
                        "location" : location.get_text(strip=True) if location else loc,
                        "stipend"  : stipend.get_text(strip=True) if stipend else "Not disclosed",
                        "duration" : duration.get_text(strip=True) if duration else "N/A",
                        "posted"   : "< 7 days",
                        "deadline" : deadline.get_text(strip=True).replace("Apply By:", "").strip() if deadline else "Not mentioned",
                        "status"   : "Not Applied",
                        "platform" : "Internshala",
                        "link"     : "https://internshala.com" + link_tag["href"] if link_tag and link_tag.get("href") else url
                    })
        except Exception as e:
            print(f"Internshala error ({loc}): {e}")
    print(f"Internshala: {len(results)} results")
    return results

# ─────────────────────────────────────────────
# SOURCE 2: INDEED
# ─────────────────────────────────────────────
def scrape_indeed():
    results = []
    queries = [
        "finance+internship", "investment+banking+intern",
        "equity+research+intern", "audit+intern",
        "wealth+management+intern", "private+equity+intern",
        "fintech+intern", "valuation+intern",
        "credit+analyst+intern", "compliance+intern",
        "treasury+intern", "financial+analyst+intern",
        "accounting+intern", "KPO+finance+intern",
    ]
    for q in queries:
        try:
            time.sleep(random.uniform(1, 2))
            url = f"https://in.indeed.com/jobs?q={q}&l=Delhi+NCR&fromage=7"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select(".job_seen_beacon")[:8]
            for job in jobs:
                title    = job.select_one('[class*="jobTitle"]')
                company  = job.select_one('[class*="companyName"]')
                location = job.select_one('[class*="companyLocation"]')
                salary   = job.select_one('[class*="salary"]')
                link_tag = job.select_one("a[id]")
                jk = link_tag["id"].replace("job_", "") if link_tag and link_tag.get("id") else ""
                t = title.get_text(strip=True) if title else "Finance Intern"
                c = company.get_text(strip=True) if company else "N/A"
                results.append({
                    "title"    : t,
                    "company"  : c,
                    "firm_type": "Corporate",
                    "domain"   : detect_domain(t, c),
                    "location" : location.get_text(strip=True) if location else "Delhi NCR",
                    "stipend"  : salary.get_text(strip=True) if salary else "Not disclosed",
                    "duration" : "3-6 Months",
                    "posted"   : "< 7 days",
                    "deadline" : "Not mentioned",
                    "status"   : "Not Applied",
                    "platform" : "Indeed",
                    "link"     : f"https://in.indeed.com/viewjob?jk={jk}" if jk else url
                })
        except Exception as e:
            print(f"Indeed error ({q}): {e}")
    print(f"Indeed: {len(results)} results")
    return results

# ─────────────────────────────────────────────
# SOURCE 3: GLASSDOOR
# ─────────────────────────────────────────────
def scrape_glassdoor():
    results = []
    searches = [
        ("finance-internship-delhi", "Finance Operations"),
        ("investment-banking-intern-delhi", "Investment Banking"),
        ("equity-research-intern-gurgaon", "Equity Research"),
        ("audit-intern-delhi", "Audit / Assurance"),
        ("private-equity-intern-delhi", "Private Equity"),
        ("fintech-intern-noida", "Fintech / Payments"),
        ("wealth-management-intern-delhi", "Wealth Management"),
        ("credit-analyst-intern-gurugram", "Credit / Debt"),
    ]
    for slug, fallback_domain in searches:
        try:
            time.sleep(random.uniform(1, 2))
            url = f"https://www.glassdoor.co.in/Job/{slug}-jobs-SRCH_KO0,{len(slug)}.htm"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select("[data-test='jobListing']")[:8]
            for job in jobs:
                title   = job.select_one("[data-test='job-title']") or job.select_one(".job-title")
                company = job.select_one("[data-test='employer-name']") or job.select_one(".employer-name")
                loc     = job.select_one("[data-test='emp-location']") or job.select_one(".location")
                link    = job.select_one("a")
                t = title.get_text(strip=True) if title else f"{fallback_domain} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                if link and link.get("href"):
                    href = "https://www.glassdoor.co.in" + link["href"] if link["href"].startswith("/") else link["href"]
                else:
                    href = url
                results.append({
                    "title"    : t,
                    "company"  : c,
                    "firm_type": "Corporate / MNC",
                    "domain"   : detect_domain(t, c) if detect_domain(t, c) != "Finance Operations" else fallback_domain,
                    "location" : loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend"  : "Not disclosed",
                    "duration" : "3-6 Months",
                    "posted"   : "< 7 days",
                    "deadline" : "Not mentioned",
                    "status"   : "Not Applied",
                    "platform" : "Glassdoor",
                    "link"     : href
                })
        except Exception as e:
            print(f"Glassdoor error ({slug}): {e}")
    print(f"Glassdoor: {len(results)} results")
    return results

# ─────────────────────────────────────────────
# SOURCE 4: NAUKRI (replaces LinkedIn)
# ─────────────────────────────────────────────
def scrape_naukri():
    results = []
    searches = [
        ("finance-internship-jobs-in-delhi-ncr", "Finance Operations"),
        ("investment-banking-internship-jobs-in-delhi-ncr", "Investment Banking"),
        ("equity-research-internship-jobs-in-delhi-ncr", "Equity Research"),
        ("audit-internship-jobs-in-delhi-ncr", "Audit / Assurance"),
        ("wealth-management-internship-jobs-in-delhi-ncr", "Wealth Management"),
        ("credit-analyst-internship-jobs-in-delhi-ncr", "Credit / Debt"),
        ("financial-analyst-internship-jobs-in-delhi-ncr", "Finance Operations"),
        ("compliance-internship-jobs-in-delhi-ncr", "Compliance / Risk"),
        ("fintech-internship-jobs-in-delhi-ncr", "Fintech / Payments"),
        ("treasury-internship-jobs-in-delhi-ncr", "Treasury"),
        ("private-equity-internship-jobs-in-delhi-ncr", "Private Equity"),
        ("valuation-internship-jobs-in-delhi-ncr", "Valuation"),
    ]
    for slug, fallback_domain in searches:
        try:
            time.sleep(random.uniform(1.5, 2.5))
            url = f"https://www.naukri.com/{slug}"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select(".jobTuple") or soup.select(".job-container") or soup.select("article.jobTupleHeader")
            for job in jobs[:8]:
                title   = job.select_one(".title") or job.select_one(".jobTitle") or job.select_one("a.title")
                company = job.select_one(".companyInfo span") or job.select_one(".company-name") or job.select_one(".subTitle")
                loc     = job.select_one(".location") or job.select_one(".locWdth")
                link    = job.select_one("a.title") or job.select_one("a[href*='naukri']")
                t = title.get_text(strip=True) if title else f"{fallback_domain} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                href = link["href"] if link and link.get("href") else url
                results.append({
                    "title"    : t,
                    "company"  : c,
                    "firm_type": "Corporate / MNC",
                    "domain"   : detect_domain(t, c) if detect_domain(t, c) != "Finance Operations" else fallback_domain,
                    "location" : loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend"  : "Not disclosed",
                    "duration" : "3-6 Months",
                    "posted"   : "< 7 days",
                    "deadline" : "Not mentioned",
                    "status"   : "Not Applied",
                    "platform" : "Naukri",
                    "link"     : href if href.startswith("http") else f"https://www.naukri.com/{slug}"
                })
        except Exception as e:
            print(f"Naukri error ({slug}): {e}")
    print(f"Naukri: {len(results)} results")
    return results

# ─────────────────────────────────────────────
# SOURCE 5: FOUNDIT (replaces LinkedIn)
# ─────────────────────────────────────────────
def scrape_foundit():
    results = []
    searches = [
        ("finance", "Finance Operations"),
        ("investment-banking", "Investment Banking"),
        ("equity-research", "Equity Research"),
        ("audit", "Audit / Assurance"),
        ("wealth-management", "Wealth Management"),
        ("fintech", "Fintech / Payments"),
        ("credit-analyst", "Credit / Debt"),
        ("compliance", "Compliance / Risk"),
    ]
    for keyword, fallback_domain in searches:
        try:
            time.sleep(random.uniform(1.5, 2.5))
            url = f"https://www.foundit.in/search/{keyword}-internship-jobs-in-delhi?experienceRanges=0~0"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select(".jobCard") or soup.select(".card-apply-content") or soup.select(".job-container")
            for job in jobs[:8]:
                title   = job.select_one(".jobTitle") or job.select_one(".designation") or job.select_one("h3")
                company = job.select_one(".companyName") or job.select_one(".company") or job.select_one("h4")
                loc     = job.select_one(".location") or job.select_one(".loc")
                link    = job.select_one("a")
                t = title.get_text(strip=True) if title else f"{fallback_domain} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                if link and link.get("href"):
                    href = "https://www.foundit.in" + link["href"] if link["href"].startswith("/") else link["href"]
                else:
                    href = url
                results.append({
                    "title"    : t,
                    "company"  : c,
                    "firm_type": "Corporate",
                    "domain"   : detect_domain(t, c) if detect_domain(t, c) != "Finance Operations" else fallback_domain,
                    "location" : loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend"  : "Not disclosed",
                    "duration" : "3-6 Months",
                    "posted"   : "< 7 days",
                    "deadline" : "Not mentioned",
                    "status"   : "Not Applied",
                    "platform" : "Foundit",
                    "link"     : href
                })
        except Exception as e:
            print(f"Foundit error ({keyword}): {e}")
    print(f"Foundit: {len(results)} results")
    return results

# ─────────────────────────────────────────────
# SOURCE 6: JOBAAJ (replaces LinkedIn)
# ─────────────────────────────────────────────
def scrape_jobaaj():
    results = []
    urls = [
        ("https://www.jobaaj.com/jobs/finance-internship-in-delhi", "Finance Operations"),
        ("https://www.jobaaj.com/jobs/investment-banking-internship-in-delhi", "Investment Banking"),
        ("https://www.jobaaj.com/jobs/equity-research-internship-in-delhi", "Equity Research"),
        ("https://www.jobaaj.com/jobs/audit-internship-in-delhi", "Audit / Assurance"),
        ("https://www.jobaaj.com/jobs/private-equity-internship-in-delhi", "Private Equity"),
        ("https://www.jobaaj.com/jobs/wealth-management-internship-in-delhi", "Wealth Management"),
        ("https://www.jobaaj.com/jobs/fintech-internship-in-delhi", "Fintech / Payments"),
        ("https://www.jobaaj.com/jobs/compliance-internship-in-delhi", "Compliance / Risk"),
    ]
    for url, fallback_domain in urls:
        try:
            time.sleep(random.uniform(1.5, 2.5))
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select(".job-list-item") or soup.select(".job-card") or soup.select(".jobcard")
            for job in jobs[:8]:
                title   = job.select_one("h2") or job.select_one(".job-title") or job.select_one("h3")
                company = job.select_one(".company") or job.select_one(".employer") or job.select_one("h4")
                loc     = job.select_one(".location") or job.select_one(".city")
                link    = job.select_one("a")
                t = title.get_text(strip=True) if title else f"{fallback_domain} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                if link and link.get("href"):
                    href = "https://www.jobaaj.com" + link["href"] if link["href"].startswith("/") else link["href"]
                else:
                    href = url
                results.append({
                    "title"    : t,
                    "company"  : c,
                    "firm_type": "Corporate / MNC",
                    "domain"   : detect_domain(t, c) if detect_domain(t, c) != "Finance Operations" else fallback_domain,
                    "location" : loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend"  : "Not disclosed",
                    "duration" : "3-6 Months",
                    "posted"   : "< 7 days",
                    "deadline" : "Not mentioned",
                    "status"   : "Not Applied",
                    "platform" : "Jobaaj",
                    "link"     : href if href.startswith("http") else url
                })
        except Exception as e:
            print(f"Jobaaj error ({url}): {e}")
    print(f"Jobaaj: {len(results)} results")
    return results

# ─────────────────────────────────────────────
# QUALITY SEED LISTINGS (always included)
# ─────────────────────────────────────────────
def get_quality_listings():
    return [
        {"title":"Finance Intern – PPO (Rs.25K/mo)","company":"TravClan Technology","firm_type":"Fintech Startup","domain":"Fintech / Payments","location":"Connaught Place, Delhi","stipend":"Rs.25,000/mo","duration":"6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Indeed","link":"https://in.indeed.com/viewjob?jk=44b725b689aa3123"},
        {"title":"Wealth Management Trainee","company":"Mint Global LLC","firm_type":"Wealth / Investment Firm","domain":"Wealth Management","location":"Delhi / Gurugram / Noida","stipend":"Rs.15,000/mo + Incentives","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Internshala","link":"https://internshala.com/internships/finance-internship-in-delhi/"},
        {"title":"Financial Audit Intern","company":"PwC India","firm_type":"Big 4 / MNC","domain":"Audit / Assurance","location":"Delhi NCR","stipend":"As per norms","duration":"Structured","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Jobaaj","link":"https://www.jobaaj.com/job/pricewaterhousecoopers-pwc-financial-audit-intern-delhi-ncr-0-to-1-years-819084"},
        {"title":"Investment Banking Intern","company":"HSBC India","firm_type":"MNC Investment Bank","domain":"Investment Banking","location":"Delhi NCR","stipend":"As per norms","duration":"Structured","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Jobaaj","link":"https://www.jobaaj.com/job/hsbc-investment-banking-intern-delhi-ncr-0-to-1-years-670612"},
        {"title":"Corporate Finance Intern","company":"Barclays","firm_type":"MNC Investment Bank","domain":"Investment Banking","location":"Delhi NCR","stipend":"As per norms","duration":"Structured","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Jobaaj","link":"https://www.jobaaj.com/job/barclays-corporate-finance-intern-delhi-ncr-0-to-1-years-760847"},
        {"title":"Equity Research Intern","company":"Trade Brains","firm_type":"Fintech / Research","domain":"Equity Research","location":"Delhi NCR / Remote","stipend":"Rs.10,000-15,000/mo","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Internshala","link":"https://internshala.com/internships/finance-internship/"},
        {"title":"Private Equity Analyst Intern","company":"Undisclosed PE Firm","firm_type":"Private Equity","domain":"Private Equity","location":"Gurugram, Haryana","stipend":"Not disclosed","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Foundit","link":"https://www.foundit.in/search/private-equity-internship-jobs-in-delhi"},
        {"title":"Venture Capital Analyst Intern","company":"Undisclosed VC Firm","firm_type":"Venture Capital","domain":"Venture Capital","location":"Delhi / Gurugram","stipend":"Not disclosed","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/venture-capital-internship-jobs-in-delhi-ncr"},
        {"title":"Credit Analyst Intern","company":"Undisclosed NBFC","firm_type":"NBFC / Lending","domain":"Credit / Debt","location":"Gurugram, Haryana","stipend":"Rs.10,000-15,000/mo","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/credit-analyst-internship-jobs-in-delhi-ncr"},
        {"title":"Treasury Intern","company":"Undisclosed MNC","firm_type":"MNC Corporate","domain":"Treasury","location":"Gurugram, Haryana","stipend":"As per norms","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/treasury-internship-jobs-in-delhi-ncr"},
        {"title":"Compliance & Risk Intern","company":"Undisclosed Financial Firm","firm_type":"Financial Services","domain":"Compliance / Risk","location":"Delhi NCR","stipend":"Not disclosed","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Foundit","link":"https://www.foundit.in/search/compliance-internship-jobs-in-delhi"},
        {"title":"KPO Financial Analyst Intern","company":"Undisclosed KPO","firm_type":"KPO Finance","domain":"KPO Finance","location":"Noida, UP","stipend":"Rs.10,000-12,000/mo","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/kpo-internship-jobs-in-noida"},
        {"title":"FP&A Intern","company":"Undisclosed MNC","firm_type":"MNC Corporate","domain":"FP&A","location":"Gurugram, Haryana","stipend":"As per norms","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Foundit","link":"https://www.foundit.in/search/fpa-internship-jobs-in-delhi"},
        {"title":"Valuation Analyst Intern","company":"Undisclosed Boutique","firm_type":"Boutique Advisory","domain":"Valuation","location":"Delhi / Gurugram","stipend":"Rs.10,000-15,000/mo","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Internshala","link":"https://internshala.com/internships/finance-internship-in-delhi/"},
        {"title":"Hedge Fund Research Intern","company":"Undisclosed HF","firm_type":"Hedge Fund","domain":"Hedge Fund","location":"Gurugram, Haryana","stipend":"Not disclosed","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Jobaaj","link":"https://www.jobaaj.com/jobs/hedge-fund-internship-in-delhi"},
        {"title":"Asset Management Intern","company":"Undisclosed AMC","firm_type":"Asset Management","domain":"Asset Management","location":"Delhi NCR","stipend":"Not disclosed","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/asset-management-internship-jobs-in-delhi-ncr"},
    ]

# ─────────────────────────────────────────────
# BUILD EXCEL
# ─────────────────────────────────────────────
def build_excel(internships, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Internships"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font   = Font(name="Arial", size=9)
    link_font   = Font(name="Arial", size=9, color="1558BB", underline="single")
    thin = Border(
        left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),  bottom=Side(style="thin", color="BFBFBF")
    )

    today = date.today().strftime("%B %d, %Y")

    ws.merge_cells("A1:N1")
    ws["A1"] = f"Finance Internships — Delhi NCR | {today} | Sources: Internshala + Indeed + Glassdoor + Naukri + Foundit + Jobaaj"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", start_color="D6E4F0")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"] = "Color legend — Red: IB | Orange: Equity/Valuation | Purple: PE/VC/HF | Green: Wealth/AM/FP&A | Blue: Fintech/Credit/Treasury/KPO | Brown: Audit/Compliance"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="7F6000")
    ws["A2"].fill = PatternFill("solid", start_color="FFF2CC")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 22

    headers = ["#", "Role / Internship Title", "Company", "Firm Type", "Domain", "Location",
               "Stipend", "Duration", "Posted", "Deadline", "Platform", "Apply Link", "Status", "Notes"]
    ws.append(headers)
    for col in range(1, 15):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[3].height = 30

    for i, job in enumerate(internships, 1):
        r = i + 3
        domain = job.get("domain", "Finance Operations")
        color  = DOMAIN_COLORS.get(domain, "404040")
        row_bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"

        values = [i, job["title"], job["company"], job["firm_type"], job["domain"],
                  job["location"], job["stipend"], job["duration"], job["posted"],
                  job["deadline"], job["platform"], job["link"], job["status"], ""]

        for col, val in enumerate(values, 1):
            cell = ws.cell(r, col, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col == 2:
                cell.font = Font(name="Arial", size=9, bold=True, color=color)
                cell.fill = PatternFill("solid", start_color=row_bg)
            elif col == 5:
                cell.font = Font(name="Arial", size=9, color=color)
                cell.fill = PatternFill("solid", start_color=row_bg)
            elif col == 12:
                cell.font = link_font
                cell.fill = PatternFill("solid", start_color=row_bg)
                cell.hyperlink = str(val)
            elif col == 13:
                cell.font = Font(name="Arial", size=9, bold=True, color="375623")
                cell.fill = PatternFill("solid", start_color="E2EFDA")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = data_font
                cell.fill = PatternFill("solid", start_color=row_bg)
        ws.row_dimensions[r].height = 36

    col_widths = [4, 40, 28, 26, 24, 26, 22, 14, 12, 18, 14, 52, 18, 22]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:N{3 + len(internships)}"

    # ── Sheet 2: Domain Summary ───────────────────────────────────────
    ws2 = wb.create_sheet("Domain Summary")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = f"Internship Count by Domain — {today}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws2["A1"].fill = PatternFill("solid", start_color="D6E4F0")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(["Domain", "Count", "Source Platforms"], 1):
        cell = ws2.cell(2, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    domain_counts = {}
    domain_platforms = {}
    for job in internships:
        d = job.get("domain", "Finance Operations")
        p = job.get("platform", "")
        domain_counts[d] = domain_counts.get(d, 0) + 1
        if d not in domain_platforms:
            domain_platforms[d] = set()
        domain_platforms[d].add(p)

    for row_idx, (domain, count) in enumerate(sorted(domain_counts.items(), key=lambda x: -x[1]), 3):
        color = DOMAIN_COLORS.get(domain, "404040")
        bg = "F5F5F5" if row_idx % 2 == 0 else "FFFFFF"
        c1 = ws2.cell(row_idx, 1, domain)
        c2 = ws2.cell(row_idx, 2, count)
        c3 = ws2.cell(row_idx, 3, ", ".join(sorted(domain_platforms.get(domain, []))))
        c1.font = Font(name="Arial", size=10, bold=True, color=color)
        c2.font = Font(name="Arial", size=10, bold=True)
        c3.font = Font(name="Arial", size=9)
        c2.alignment = Alignment(horizontal="center")
        for c in [c1, c2, c3]:
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin

    last = len(domain_counts) + 3
    for col, val in enumerate(["TOTAL", f"=SUM(B3:B{last-1})", ""], 1):
        cell = ws2.cell(last, col, val)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill("solid", start_color="D6E4F0")
        cell.border = thin
        if col == 2:
            cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 40

    # ── Sheet 3: Application Tracker ─────────────────────────────────
    ws3 = wb.create_sheet("Application Tracker")
    ws3.merge_cells("A1:H1")
    ws3["A1"] = f"My Application Tracker — {today}"
    ws3["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws3["A1"].fill = PatternFill("solid", start_color="D6E4F0")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    for col, h in enumerate(["#", "Company", "Role", "Domain", "Applied Date", "Status", "Interview Date", "Notes"], 1):
        cell = ws3.cell(2, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws3.row_dimensions[2].height = 28

    status_map = [
        ("Not Applied", "F2F2F2"), ("Applied", "BDD7EE"),
        ("In Review",   "FFF2CC"), ("Interview Scheduled", "E2EFDA"),
        ("Offer Received", "C6EFCE"), ("Rejected", "FFE0E0"),
    ]
    for i in range(3, 53):
        row_bg = "FFFFFF" if i % 2 == 0 else "F9F9F9"
        for col in range(1, 9):
            cell = ws3.cell(i, col)
            cell.fill = PatternFill("solid", start_color=row_bg)
            cell.border = thin
            cell.font = data_font
            if col == 1:
                cell.value = i - 2
                cell.alignment = Alignment(horizontal="center")

    ws3.merge_cells("A55:H55")
    ws3["A55"] = "STATUS LEGEND"
    ws3["A55"].font = Font(name="Arial", bold=True, size=10)
    ws3["A55"].fill = PatternFill("solid", start_color="1F4E79")
    ws3["A55"].font = Font(name="Arial", bold=True, color="FFFFFF")
    ws3["A55"].alignment = Alignment(horizontal="center")

    for idx, (status, color) in enumerate(status_map, 56):
        cell = ws3.cell(idx, 1, status)
        cell.font = Font(name="Arial", bold=True, size=9)
        cell.fill = PatternFill("solid", start_color=color)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")
        ws3.merge_cells(f"B{idx}:H{idx}")
        ws3.cell(idx, 2).fill = PatternFill("solid", start_color=color)
        ws3.cell(idx, 2).border = thin

    for col, w in zip(range(1, 9), [4, 28, 36, 24, 16, 22, 18, 36]):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A3"

    wb.save(filepath)
    print(f"Excel saved: {filepath}")

# ─────────────────────────────────────────────
# SEND EMAIL
# ─────────────────────────────────────────────
def send_email(filepath, internships):
    today_str = date.today().strftime("%B %d, %Y")
    count = len(internships)

    domain_counts = {}
    platform_counts = {}
    for job in internships:
        d = job.get("domain", "Finance Operations")
        p = job.get("platform", "Other")
        domain_counts[d] = domain_counts.get(d, 0) + 1
        platform_counts[p] = platform_counts.get(p, 0) + 1

    domain_lines   = "\n".join([f"   {d}: {c}" for d, c in sorted(domain_counts.items(), key=lambda x: -x[1])])
    platform_lines = "\n".join([f"   {p}: {c}" for p, c in sorted(platform_counts.items(), key=lambda x: -x[1])])

    msg = MIMEMultipart()
    msg["From"]    = YOUR_EMAIL
    msg["To"]      = SEND_TO_EMAIL
    msg["Subject"] = f"Finance Internships Delhi NCR — {today_str} — {count} Listings"

    body = f"""Hi,

Your daily Finance Internship tracker (v3.0) is ready!

DATE     : {today_str}
TOTAL    : {count} internships
LOCATION : Delhi / Gurugram / Noida / Greater Noida

SOURCES SCRAPED:
{platform_lines}

DOMAIN BREAKDOWN:
{domain_lines}

EXCEL HAS 3 SHEETS:
  Sheet 1 — All Internships (color coded + apply links + status column)
  Sheet 2 — Domain Summary (count per domain + which platforms)
  Sheet 3 — Application Tracker (50 rows to track your applications)

HOW TO USE:
  1. Open Excel
  2. Filter by Domain column to find your preferred area
  3. Click Apply Link to go directly to the listing
  4. Update STATUS column after applying
  5. Use Sheet 3 to track interview progress

Good luck!
-- Internship Bot v3.0 (GitHub Actions)
   Sources: Internshala + Indeed + Glassdoor + Naukri + Foundit + Jobaaj
"""
    msg.attach(MIMEText(body, "plain"))
    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")
        msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(YOUR_EMAIL, YOUR_PASSWORD)
        server.sendmail(YOUR_EMAIL, SEND_TO_EMAIL, msg.as_string())
    print("Email sent!")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def run():
    print("Starting Internship Bot v3.0...")
    all_jobs = []
    all_jobs += scrape_internshala()
    all_jobs += scrape_indeed()
    all_jobs += scrape_glassdoor()
    all_jobs += scrape_naukri()
    all_jobs += scrape_foundit()
    all_jobs += scrape_jobaaj()
    all_jobs += get_quality_listings()

    seen, unique = set(), []
    for j in all_jobs:
        key = (j["title"].strip().lower()[:40], j["company"].strip().lower()[:30])
        if key not in seen:
            seen.add(key)
            unique.append(j)

    unique.sort(key=lambda x: x.get("domain", "Z"))

    today_str = date.today().strftime("%Y-%m-%d")
    filepath = f"/tmp/Finance_Internships_DelhiNCR_{today_str}.xlsx"
    build_excel(unique, filepath)
    send_email(filepath, unique)
    print(f"Done! {len(unique)} internships sent.")

if __name__ == "__main__":
    run()

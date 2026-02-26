import smtplib, os, time, random, json
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup

# ── Google Sheets (gspread) ──────────────────
import gspread
from google.oauth2.service_account import Credentials

# ── Anthropic Claude API ─────────────────────
import anthropic

# ─────────────────────────────────────────────
# ENV / CONFIG
# ─────────────────────────────────────────────
YOUR_EMAIL    = os.environ["EMAIL"]
YOUR_PASSWORD = os.environ["PASSWORD"]
SEND_TO_EMAIL = os.environ["EMAIL"]
ANTHROPIC_KEY = os.environ["ANTHROPIC_API_KEY"]

# Google Sheets: service account JSON stored as GitHub Secret
# Set secret name: GOOGLE_SERVICE_ACCOUNT_JSON  (full JSON string)
# Set secret name: GSHEET_ID  (the spreadsheet ID from the URL)
GSHEET_ID     = os.environ["GSHEET_ID"]
GSHEET_TAB    = "seen_listings"   # tab name for deduplication log

# Adzuna API credentials
ADZUNA_APP_ID  = os.environ.get("ADZUNA_APP_ID",  "d69dc243")
ADZUNA_APP_KEY = os.environ.get("ADZUNA_APP_KEY", "d0727414dbf421e1c5b42f97019bb50d")

# ─────────────────────────────────────────────
# SCHEDULE: Mon, Thu, Sat, Sun
# ─────────────────────────────────────────────
RUN_ON_WEEKDAYS = {0, 1, 2, 3, 4, 5, 6}  # Every day

# ─────────────────────────────────────────────
# TANISHQ'S RESUME (for AI fit scoring)
# ─────────────────────────────────────────────
RESUME_TEXT = """
Name: Tanishq Singhal
Degree: BBA Finance & Banking — IMS UCC Ghaziabad (2024–2027)
Certification: SEBI NISM Investor Certification (valid till March 2027)

Experience:
- Founder Fellow @ 23 Ventures (Dec 2025–Feb 2026): Secured startup credits (Cloudflare, OpenAI,
  Mixpanel), hosted offline hackathon (14 teams, 30+ participants), partnered with 25+ colleges
  across Delhi NCR and Pune for hackathons and webinars.
- Finance Intern @ Yhills, Noida (Jun–Aug 2025): Built financial models enabling 30% faster
  evaluations, prepared 5+ financial reports with company profiles and forecasts, conducted
  revenue/cost trend analysis using Advanced Excel and Google Sheets.
- Campus Ambassador @ Yhills (May–Jun 2025): Generated 40+ registrations (100%+ growth),
  coordinated 5+ events, managed 40+ member student community.

Skills:
- Technical: Advanced Excel (VLOOKUP, Index-Match, Pivot), Power BI, Google Sheets
- Finance: Financial Modeling, Ratio Analysis, EBITDA, Forecasting, Sales Growth
- Soft: Leadership, Event Management, Stakeholder Communication, Problem-Solving

Achievements:
- Winter Consulting 2025 – IIT Guwahati (Top 10%: Consulting & Strategy)
- McKinsey Forward Program (Dec 2025: Problem-Solving & Strategic Thinking)

Best-fit domains: Founder's Office, Strategy & Consulting, Investment Banking,
Equity Research, Valuation, FP&A, Operations / Growth, Management Trainee
"""

HEADERS_LIST = [
    {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"},
    {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/119.0.0.0 Safari/537.36"},
    {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/118.0.0.0 Safari/537.36"},
]

DOMAIN_COLORS = {
    "Investment Banking":   "C00000",
    "Equity Research":      "E26B0A",
    "Wealth Management":    "375623",
    "Private Equity":       "7030A0",
    "Venture Capital":      "7030A0",
    "Hedge Fund":           "7030A0",
    "Credit / Debt":        "1F4E79",
    "Treasury":             "1F4E79",
    "Compliance / Risk":    "833C00",
    "Audit / Assurance":    "833C00",
    "Valuation":            "E26B0A",
    "FP&A":                 "375623",
    "Fintech / Payments":   "0070C0",
    "Finance Operations":   "404040",
    "Accounting / Tax":     "404040",
    "Asset Management":     "375623",
    "Portfolio Management": "375623",
    "KPO Finance":          "1F4E79",
    "Founder's Office":     "A50021",
    "Management Trainee":   "BF8F00",
    "Strategy & Consulting":"1D6A96",
    "Operations / Growth":  "4A7C59",
}

def get_headers():
    return random.choice(HEADERS_LIST)

def detect_domain(title, company=""):
    text = (title + " " + company).lower()
    if any(x in text for x in ["founder", "chief of staff", "cxo office", "ceo office", "founders office"]):
        return "Founder's Office"
    if any(x in text for x in ["management trainee", "mt program", "graduate trainee", "management intern"]):
        return "Management Trainee"
    if any(x in text for x in ["strategy", "consulting intern", "management consulting", "business consulting", "mckinsey", "bcg", "bain", "deloitte", "kpmg consulting"]):
        return "Strategy & Consulting"
    if any(x in text for x in ["operations intern", "growth intern", "growth hacking", "business operations", "ops intern", "revenue operations"]):
        return "Operations / Growth"
    if any(x in text for x in ["investment bank", "ib ", "m&a", "capital market", "corporate finance"]): return "Investment Banking"
    if any(x in text for x in ["equity research", "equity analyst", "stock analyst", "market research", "research analyst"]): return "Equity Research"
    if any(x in text for x in ["wealth", "private banking", "relationship manager", "hni"]): return "Wealth Management"
    if any(x in text for x in ["private equity", " pe ", "buyout", "growth equity"]): return "Private Equity"
    if any(x in text for x in ["venture", " vc ", "startup invest"]): return "Venture Capital"
    if any(x in text for x in ["hedge", "quant", "algo", "derivatives"]): return "Hedge Fund"
    if any(x in text for x in ["credit", "debt", "lending", "fixed income", "bond", "nbfc"]): return "Credit / Debt"
    if any(x in text for x in ["treasury", "forex", " fx ", "currency", "liquidity"]): return "Treasury"
    if any(x in text for x in ["compliance", " risk", "regulatory", "kyc", "aml"]): return "Compliance / Risk"
    if any(x in text for x in ["audit", "assurance", "internal audit"]): return "Audit / Assurance"
    if any(x in text for x in ["valuat", "dcf", "modell", "financial model"]): return "Valuation"
    if any(x in text for x in ["fp&a", "financial planning", "budgeting", "forecasting"]): return "FP&A"
    if any(x in text for x in ["fintech", "payment", "wallet", "insurtech", "neobank"]): return "Fintech / Payments"
    if any(x in text for x in ["kpo", "research process"]): return "KPO Finance"
    if any(x in text for x in ["portfolio", "asset manag", "fund manag", "aum", "mutual fund"]): return "Asset Management"
    if any(x in text for x in ["account", " tax", "gst", "tally", "bookkeep"]): return "Accounting / Tax"
    return "Finance Operations"

# ═════════════════════════════════════════════
# GOOGLE SHEETS — DEDUPLICATION ACROSS RUNS
# ═════════════════════════════════════════════

def get_gsheet_client():
    """Authenticate with Google Sheets using service account JSON from env."""
    sa_json = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    sa_info = json.loads(sa_json)
    scopes  = ["https://www.googleapis.com/auth/spreadsheets"]
    creds   = Credentials.from_service_account_info(sa_info, scopes=scopes)
    return gspread.authorize(creds)

def load_seen_keys(client):
    """Load all previously seen (title, company) keys from Google Sheet."""
    try:
        sh  = client.open_by_key(GSHEET_ID)
        try:
            ws = sh.worksheet(GSHEET_TAB)
        except gspread.exceptions.WorksheetNotFound:
            # First run — create the tab with headers
            ws = sh.add_worksheet(title=GSHEET_TAB, rows=5000, cols=4)
            ws.append_row(["title_key", "company_key", "domain", "first_seen"])
        rows = ws.get_all_values()
        seen = set()
        for row in rows[1:]:   # skip header
            if len(row) >= 2:
                seen.add((row[0], row[1]))
        print(f"Loaded {len(seen)} previously seen listings from Google Sheets.")
        return seen, ws
    except Exception as e:
        print(f"Google Sheets load error: {e}")
        return set(), None

def save_new_keys(ws, new_jobs):
    """Append newly seen jobs to the Google Sheet dedup log."""
    if not ws or not new_jobs:
        return
    try:
        today_str = date.today().isoformat()
        rows = [
            [j["title"].strip().lower()[:40],
             j["company"].strip().lower()[:30],
             j.get("domain", ""),
             today_str]
            for j in new_jobs
        ]
        ws.append_rows(rows, value_input_option="RAW")
        print(f"Saved {len(rows)} new listings to Google Sheets dedup log.")
    except Exception as e:
        print(f"Google Sheets save error: {e}")

def filter_new_only(all_jobs, seen_keys):
    """Return only jobs not seen in previous runs."""
    new_jobs, dup_count = [], 0
    local_seen = set()
    for j in all_jobs:
        key = (j["title"].strip().lower()[:40], j["company"].strip().lower()[:30])
        if key in seen_keys or key in local_seen:
            dup_count += 1
        else:
            local_seen.add(key)
            new_jobs.append(j)
    print(f"Dedup: {dup_count} duplicates removed, {len(new_jobs)} new listings kept.")
    return new_jobs

# ═════════════════════════════════════════════
# AI FIT SCORING — CLAUDE API
# ═════════════════════════════════════════════

def score_jobs_with_ai(jobs):
    """
    Batch score all jobs using Claude.
    Returns list of jobs with 'fit_score' (1-10) and 'fit_reason' (short string) added.
    To save API calls, we send all jobs in one prompt and parse JSON back.
    """
    if not jobs:
        return jobs

    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

    # Build a compact job list for the prompt
    job_summaries = []
    for i, j in enumerate(jobs):
        job_summaries.append(f"{i}: {j['title']} @ {j['company']} | Domain: {j['domain']} | Location: {j['location']} | Stipend: {j['stipend']}")
    job_text = "\n".join(job_summaries)

    prompt = f"""You are a career advisor. Score each internship listing for fit with the candidate below.

CANDIDATE RESUME:
{RESUME_TEXT}

INTERNSHIP LISTINGS (index: title @ company | domain | location | stipend):
{job_text}

For each listing, return a JSON array (same order, same indexes) like:
[
  {{"index": 0, "fit_score": 8, "fit_reason": "Matches finance modeling skills and SEBI cert"}},
  {{"index": 1, "fit_score": 5, "fit_reason": "Partial match, accounting-heavy role"}},
  ...
]

Scoring guide:
- 9-10: Perfect match (domain aligns, uses candidate's exact skills, good for BBA 1st year)
- 7-8:  Strong match (related domain, transferable skills)
- 5-6:  Moderate match (some overlap but gaps exist)
- 3-4:  Weak match (different domain, limited overlap)
- 1-2:  Poor match (irrelevant to profile)

Return ONLY the JSON array, no explanation, no markdown."""

    try:
        response = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response.content[0].text.strip()
        # Strip markdown fences if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        scores = json.loads(raw.strip())
        score_map = {item["index"]: item for item in scores}
        for i, job in enumerate(jobs):
            info = score_map.get(i, {})
            job["fit_score"]  = info.get("fit_score", 5)
            job["fit_reason"] = info.get("fit_reason", "Not scored")
        print(f"AI fit scoring complete for {len(jobs)} listings.")
    except Exception as e:
        print(f"AI scoring error: {e}")
        for job in jobs:
            job.setdefault("fit_score", 5)
            job.setdefault("fit_reason", "Scoring unavailable")

    return jobs

# ═════════════════════════════════════════════
# SOURCE 1: INTERNSHALA
# ═════════════════════════════════════════════
def scrape_internshala():
    results = []
    urls = [
        ("https://internshala.com/internships/finance-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/finance-internship-in-gurgaon/", "Gurugram"),
        ("https://internshala.com/internships/finance-internship-in-noida/", "Noida"),
        ("https://internshala.com/internships/investment-banking-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/equity-research-internship/", "Delhi NCR"),
        ("https://internshala.com/internships/audit-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/mba-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/accounting-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/strategy-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/operations-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/management-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/business-development-internship-in-delhi/", "Delhi"),
        ("https://internshala.com/internships/general-management-internship/", "Delhi NCR"),
        ("https://internshala.com/internships/consulting-internship-in-delhi/", "Delhi"),
    ]
    for url, loc in urls:
        try:
            time.sleep(random.uniform(1, 2))
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            # Each internship container — try multiple selectors
            cards = (soup.select(".internship_meta") or
                     soup.select(".internship-list-item") or
                     soup.select("[id^=internship_]"))[:20]
            for card in cards:
                title    = card.select_one(".profile") or card.select_one(".job-title")
                company  = card.select_one(".company_name") or card.select_one(".company-name")
                location = card.select_one(".location_link") or card.select_one(".locations")
                stipend  = card.select_one(".stipend") or card.select_one(".stipend-container")
                duration = card.select_one(".item_body") or card.select_one(".duration-container")
                deadline = card.select_one(".apply-by") or card.select_one(".close-by")

                # ── Direct job link extraction ──
                # Internshala job URLs follow pattern: /internship/detail/SLUG
                direct_link = None
                # Try: href on the title anchor
                title_a = card.select_one("a.job-title-href") or card.select_one("a[href*='/internship/detail/']")
                if title_a and title_a.get("href"):
                    direct_link = "https://internshala.com" + title_a["href"]
                # Try: parent anchor of the card
                if not direct_link:
                    parent_a = card.find_parent("a")
                    if parent_a and parent_a.get("href") and "/internship/" in parent_a["href"]:
                        direct_link = "https://internshala.com" + parent_a["href"]
                # Try: any anchor inside card pointing to internship detail
                if not direct_link:
                    any_a = card.select_one("a[href*='/internship/']")
                    if any_a and any_a.get("href"):
                        href = any_a["href"]
                        direct_link = "https://internshala.com" + href if href.startswith("/") else href
                # Final fallback: search page (at least narrows to category)
                if not direct_link:
                    direct_link = url

                if title:
                    t = title.get_text(strip=True)
                    c = company.get_text(strip=True) if company else "N/A"
                    results.append({
                        "title":     t, "company": c,
                        "firm_type": "Startup / Corporate",
                        "domain":    detect_domain(t, c),
                        "location":  location.get_text(strip=True) if location else loc,
                        "stipend":   stipend.get_text(strip=True) if stipend else "Not disclosed",
                        "duration":  duration.get_text(strip=True) if duration else "N/A",
                        "posted":    "< 7 days",
                        "deadline":  deadline.get_text(strip=True).replace("Apply By:", "").strip() if deadline else "Not mentioned",
                        "status":    "Not Applied", "platform": "Internshala",
                        "link":      direct_link
                    })
        except Exception as e:
            print(f"Internshala error ({loc}): {e}")
    print(f"Internshala: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# SOURCE 2: INDEED
# ═════════════════════════════════════════════
def scrape_indeed():
    results = []
    queries = [
        "finance+internship", "investment+banking+intern", "equity+research+intern",
        "audit+intern", "wealth+management+intern", "private+equity+intern",
        "fintech+intern", "valuation+intern", "credit+analyst+intern",
        "compliance+intern", "treasury+intern", "financial+analyst+intern",
        "accounting+intern", "KPO+finance+intern", "founder+office+intern",
        "chief+of+staff+intern", "management+trainee+delhi", "strategy+intern",
        "consulting+intern", "operations+intern", "growth+intern",
    ]
    for q in queries:
        try:
            time.sleep(random.uniform(1, 2))
            url = f"https://in.indeed.com/jobs?q={q}&l=Delhi+NCR&fromage=7"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            for job in soup.select(".job_seen_beacon")[:8]:
                title    = job.select_one('[class*="jobTitle"]')
                company  = job.select_one('[class*="companyName"]')
                location = job.select_one('[class*="companyLocation"]')
                salary   = job.select_one('[class*="salary"]')
                link_tag = job.select_one("a[id]")
                jk = link_tag["id"].replace("job_", "") if link_tag and link_tag.get("id") else ""
                t = title.get_text(strip=True) if title else "Intern"
                c = company.get_text(strip=True) if company else "N/A"
                results.append({
                    "title": t, "company": c, "firm_type": "Corporate",
                    "domain": detect_domain(t, c),
                    "location": location.get_text(strip=True) if location else "Delhi NCR",
                    "stipend": salary.get_text(strip=True) if salary else "Not disclosed",
                    "duration": "3-6 Months", "posted": "< 7 days",
                    "deadline": "Not mentioned", "status": "Not Applied",
                    "platform": "Indeed",
                    "link": f"https://in.indeed.com/viewjob?jk={jk}" if jk else url
                })
        except Exception as e:
            print(f"Indeed error ({q}): {e}")
    print(f"Indeed: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# SOURCE 3: GLASSDOOR
# ═════════════════════════════════════════════
def scrape_glassdoor():
    results = []
    searches = [
        ("finance-internship-delhi", "Finance Operations"),
        ("investment-banking-intern-delhi", "Investment Banking"),
        ("equity-research-intern-gurgaon", "Equity Research"),
        ("audit-intern-delhi", "Audit / Assurance"),
        ("private-equity-intern-delhi", "Private Equity"),
        ("fintech-intern-noida", "Fintech / Payments"),
        ("wealth-management-intern-delhi", "Wealth Management"),
        ("credit-analyst-intern-gurugram", "Credit / Debt"),
        ("founders-office-intern-delhi", "Founder's Office"),
        ("management-trainee-delhi", "Management Trainee"),
        ("strategy-intern-delhi", "Strategy & Consulting"),
        ("consulting-intern-delhi", "Strategy & Consulting"),
        ("operations-intern-delhi", "Operations / Growth"),
    ]
    for slug, fallback in searches:
        try:
            time.sleep(random.uniform(1, 2))
            url = f"https://www.glassdoor.co.in/Job/{slug}-jobs-SRCH_KO0,{len(slug)}.htm"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            for job in soup.select("[data-test='jobListing']")[:8]:
                title   = job.select_one("[data-test='job-title']") or job.select_one(".job-title")
                company = job.select_one("[data-test='employer-name']") or job.select_one(".employer-name")
                loc     = job.select_one("[data-test='emp-location']") or job.select_one(".location")
                link    = job.select_one("a")
                t = title.get_text(strip=True) if title else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                # Glassdoor direct link — job URLs contain /job-listing/ or /partner/
                direct_link = url  # fallback
                job_link = (job.select_one("a[href*='/job-listing/']") or
                            job.select_one("a[href*='/partner/']") or
                            job.select_one("a[data-test='job-title']") or
                            job.select_one("a[class*='JobCard']"))
                if job_link and job_link.get("href"):
                    h = job_link["href"]
                    direct_link = h if h.startswith("http") else "https://www.glassdoor.co.in" + h
                elif link and link.get("href"):
                    h = link["href"]
                    direct_link = h if h.startswith("http") else "https://www.glassdoor.co.in" + h
                d = detect_domain(t, c)
                results.append({
                    "title": t, "company": c, "firm_type": "Corporate / MNC",
                    "domain": d if d != "Finance Operations" else fallback,
                    "location": loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend": "Not disclosed", "duration": "3-6 Months",
                    "posted": "< 7 days", "deadline": "Not mentioned",
                    "status": "Not Applied", "platform": "Glassdoor", "link": direct_link
                })
        except Exception as e:
            print(f"Glassdoor error ({slug}): {e}")
    print(f"Glassdoor: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# SOURCE 4: NAUKRI
# ═════════════════════════════════════════════
def scrape_naukri():
    results = []
    searches = [
        ("finance-internship-jobs-in-delhi-ncr", "Finance Operations"),
        ("investment-banking-internship-jobs-in-delhi-ncr", "Investment Banking"),
        ("equity-research-internship-jobs-in-delhi-ncr", "Equity Research"),
        ("audit-internship-jobs-in-delhi-ncr", "Audit / Assurance"),
        ("wealth-management-internship-jobs-in-delhi-ncr", "Wealth Management"),
        ("credit-analyst-internship-jobs-in-delhi-ncr", "Credit / Debt"),
        ("financial-analyst-internship-jobs-in-delhi-ncr", "Finance Operations"),
        ("compliance-internship-jobs-in-delhi-ncr", "Compliance / Risk"),
        ("fintech-internship-jobs-in-delhi-ncr", "Fintech / Payments"),
        ("treasury-internship-jobs-in-delhi-ncr", "Treasury"),
        ("private-equity-internship-jobs-in-delhi-ncr", "Private Equity"),
        ("valuation-internship-jobs-in-delhi-ncr", "Valuation"),
        ("founders-office-internship-jobs-in-delhi-ncr", "Founder's Office"),
        ("management-trainee-jobs-in-delhi-ncr", "Management Trainee"),
        ("strategy-internship-jobs-in-delhi-ncr", "Strategy & Consulting"),
        ("consulting-internship-jobs-in-delhi-ncr", "Strategy & Consulting"),
        ("operations-internship-jobs-in-delhi-ncr", "Operations / Growth"),
        ("growth-internship-jobs-in-delhi-ncr", "Operations / Growth"),
    ]
    for slug, fallback in searches:
        try:
            time.sleep(random.uniform(1.5, 2.5))
            url = f"https://www.naukri.com/{slug}"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select(".jobTuple") or soup.select(".job-container") or soup.select("article.jobTupleHeader")
            for job in jobs[:8]:
                title   = job.select_one(".title") or job.select_one(".jobTitle") or job.select_one("a.title")
                company = job.select_one(".companyInfo span") or job.select_one(".company-name") or job.select_one(".subTitle")
                loc     = job.select_one(".location") or job.select_one(".locWdth")
                link    = job.select_one("a.title") or job.select_one("a[href*='naukri']")
                t = title.get_text(strip=True) if title else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                # Naukri direct link — job URLs contain /job-listings/
                direct_link = url  # fallback
                if link and link.get("href"):
                    h = link["href"]
                    if "naukri.com" in h or h.startswith("/"):
                        direct_link = h if h.startswith("http") else "https://www.naukri.com" + h
                # Also try href on title element itself
                if direct_link == url and title:
                    title_a = title if title.name == "a" else title.find("a")
                    if title_a and title_a.get("href"):
                        h = title_a["href"]
                        direct_link = h if h.startswith("http") else "https://www.naukri.com" + h
                d = detect_domain(t, c)
                results.append({
                    "title": t, "company": c, "firm_type": "Corporate / MNC",
                    "domain": d if d != "Finance Operations" else fallback,
                    "location": loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend": "Not disclosed", "duration": "3-6 Months",
                    "posted": "< 7 days", "deadline": "Not mentioned",
                    "status": "Not Applied", "platform": "Naukri",
                    "link": direct_link
                })
        except Exception as e:
            print(f"Naukri error ({slug}): {e}")
    print(f"Naukri: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# SOURCE 5: FOUNDIT
# ═════════════════════════════════════════════
def scrape_foundit():
    results = []
    searches = [
        ("finance", "Finance Operations"), ("investment-banking", "Investment Banking"),
        ("equity-research", "Equity Research"), ("audit", "Audit / Assurance"),
        ("wealth-management", "Wealth Management"), ("fintech", "Fintech / Payments"),
        ("credit-analyst", "Credit / Debt"), ("compliance", "Compliance / Risk"),
        ("founders-office", "Founder's Office"), ("management-trainee", "Management Trainee"),
        ("strategy-consulting", "Strategy & Consulting"), ("business-operations", "Operations / Growth"),
        ("growth", "Operations / Growth"),
    ]
    for keyword, fallback in searches:
        try:
            time.sleep(random.uniform(1.5, 2.5))
            url = f"https://www.foundit.in/search/{keyword}-internship-jobs-in-delhi?experienceRanges=0~0"
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select(".jobCard") or soup.select(".card-apply-content") or soup.select(".job-container")
            for job in jobs[:8]:
                title   = job.select_one(".jobTitle") or job.select_one(".designation") or job.select_one("h3")
                company = job.select_one(".companyName") or job.select_one(".company") or job.select_one("h4")
                loc     = job.select_one(".location") or job.select_one(".loc")
                link    = job.select_one("a")
                t = title.get_text(strip=True) if title else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                # Foundit direct link — job URLs contain /j/ or /job/
                direct_link = url  # fallback
                job_link = (job.select_one("a[href*='/j/']") or
                            job.select_one("a[href*='/job/']") or
                            job.select_one("a.jobTitle") or
                            job.select_one("a[href*='foundit']"))
                if job_link and job_link.get("href"):
                    h = job_link["href"]
                    direct_link = h if h.startswith("http") else "https://www.foundit.in" + h
                elif link and link.get("href"):
                    h = link["href"]
                    direct_link = h if h.startswith("http") else "https://www.foundit.in" + h
                d = detect_domain(t, c)
                results.append({
                    "title": t, "company": c, "firm_type": "Corporate",
                    "domain": d if d != "Finance Operations" else fallback,
                    "location": loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend": "Not disclosed", "duration": "3-6 Months",
                    "posted": "< 7 days", "deadline": "Not mentioned",
                    "status": "Not Applied", "platform": "Foundit", "link": direct_link
                })
        except Exception as e:
            print(f"Foundit error ({keyword}): {e}")
    print(f"Foundit: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# SOURCE 6: JOBAAJ
# ═════════════════════════════════════════════
def scrape_jobaaj():
    results = []
    urls = [
        ("https://www.jobaaj.com/jobs/finance-internship-in-delhi", "Finance Operations"),
        ("https://www.jobaaj.com/jobs/investment-banking-internship-in-delhi", "Investment Banking"),
        ("https://www.jobaaj.com/jobs/equity-research-internship-in-delhi", "Equity Research"),
        ("https://www.jobaaj.com/jobs/audit-internship-in-delhi", "Audit / Assurance"),
        ("https://www.jobaaj.com/jobs/private-equity-internship-in-delhi", "Private Equity"),
        ("https://www.jobaaj.com/jobs/wealth-management-internship-in-delhi", "Wealth Management"),
        ("https://www.jobaaj.com/jobs/fintech-internship-in-delhi", "Fintech / Payments"),
        ("https://www.jobaaj.com/jobs/compliance-internship-in-delhi", "Compliance / Risk"),
        ("https://www.jobaaj.com/jobs/founders-office-internship-in-delhi", "Founder's Office"),
        ("https://www.jobaaj.com/jobs/management-trainee-in-delhi", "Management Trainee"),
        ("https://www.jobaaj.com/jobs/strategy-internship-in-delhi", "Strategy & Consulting"),
        ("https://www.jobaaj.com/jobs/operations-internship-in-delhi", "Operations / Growth"),
    ]
    for url, fallback in urls:
        try:
            time.sleep(random.uniform(1.5, 2.5))
            r = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = soup.select(".job-list-item") or soup.select(".job-card") or soup.select(".jobcard")
            for job in jobs[:8]:
                title   = job.select_one("h2") or job.select_one(".job-title") or job.select_one("h3")
                company = job.select_one(".company") or job.select_one(".employer") or job.select_one("h4")
                loc     = job.select_one(".location") or job.select_one(".city")
                link    = job.select_one("a")
                t = title.get_text(strip=True) if title else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                # Jobaaj direct link — job URLs contain /job/
                direct_link = url  # fallback
                job_link = (job.select_one("a[href*='/job/']") or
                            job.select_one("a.job-title") or
                            job.select_one("h2 > a") or
                            job.select_one("h3 > a"))
                if job_link and job_link.get("href"):
                    h = job_link["href"]
                    direct_link = h if h.startswith("http") else "https://www.jobaaj.com" + h
                elif link and link.get("href"):
                    h = link["href"]
                    direct_link = h if h.startswith("http") else "https://www.jobaaj.com" + h
                d = detect_domain(t, c)
                results.append({
                    "title": t, "company": c, "firm_type": "Corporate / MNC",
                    "domain": d if d != "Finance Operations" else fallback,
                    "location": loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend": "Not disclosed", "duration": "3-6 Months",
                    "posted": "< 7 days", "deadline": "Not mentioned",
                    "status": "Not Applied", "platform": "Jobaaj",
                    "link": direct_link
                })
        except Exception as e:
            print(f"Jobaaj error ({url}): {e}")
    print(f"Jobaaj: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# SOURCE 7: WELLFOUND (AngelList) — NEW
# ═════════════════════════════════════════════
def scrape_wellfound():
    """
    Scrapes Wellfound (formerly AngelList Talent) for startup internships.
    Best source for Founder's Office, Strategy, and early-stage roles.
    """
    results = []
    searches = [
        ("finance", "Finance Operations"),
        ("investment", "Investment Banking"),
        ("strategy", "Strategy & Consulting"),
        ("operations", "Operations / Growth"),
        ("founder", "Founder's Office"),
        ("growth", "Operations / Growth"),
        ("business-development", "Operations / Growth"),
        ("consulting", "Strategy & Consulting"),
        ("fintech", "Fintech / Payments"),
        ("venture-capital", "Venture Capital"),
    ]
    for keyword, fallback in searches:
        try:
            time.sleep(random.uniform(2, 3))
            url = f"https://wellfound.com/role/l/{keyword}-intern/india"
            r = requests.get(url, headers=get_headers(), timeout=20)
            soup = BeautifulSoup(r.text, "html.parser")

            # Wellfound job cards
            jobs = (soup.select("[data-test='StartupResult']") or
                    soup.select(".styles_component__Ey28k") or
                    soup.select("div[class*='JobListing']") or
                    soup.select("div[class*='job-listing']"))

            for job in jobs[:10]:
                title   = (job.select_one("a[class*='JobTitle']") or
                           job.select_one("[class*='title']") or
                           job.select_one("h2") or job.select_one("h3"))
                company = (job.select_one("[class*='startup-link']") or
                           job.select_one("[class*='company']") or
                           job.select_one("h4"))
                loc     = (job.select_one("[class*='location']") or
                           job.select_one("[class*='Location']"))
                comp    = (job.select_one("[class*='compensation']") or
                           job.select_one("[class*='salary']"))
                link    = job.select_one("a[href*='/jobs/']") or job.select_one("a")

                t = title.get_text(strip=True) if title else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "Startup"

                # Build absolute href
                # Wellfound direct link — job URLs contain /jobs/ or /l/
                direct_link = url
                job_link = (job.select_one("a[href*='/jobs/']") or
                            job.select_one("a[href*='/job/']") or
                            job.select_one("a[class*='JobTitle']") or
                            link)
                if job_link and job_link.get("href"):
                    raw_href = job_link["href"]
                    direct_link = ("https://wellfound.com" + raw_href
                                   if raw_href.startswith("/") else raw_href)
                href = direct_link

                d = detect_domain(t, c)
                results.append({
                    "title":     t,
                    "company":   c,
                    "firm_type": "Startup (Wellfound)",
                    "domain":    d if d != "Finance Operations" else fallback,
                    "location":  loc.get_text(strip=True) if loc else "Delhi NCR / Remote",
                    "stipend":   comp.get_text(strip=True) if comp else "Not disclosed",
                    "duration":  "3-6 Months",
                    "posted":    "< 7 days",
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "Wellfound",
                    "link":      href
                })
        except Exception as e:
            print(f"Wellfound error ({keyword}): {e}")

    print(f"Wellfound: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# QUALITY SEED LISTINGS
# ═════════════════════════════════════════════
def get_quality_listings():
    return [
        {"title":"Finance Intern – PPO (Rs.25K/mo)","company":"TravClan Technology","firm_type":"Fintech Startup","domain":"Fintech / Payments","location":"Connaught Place, Delhi","stipend":"Rs.25,000/mo","duration":"6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Indeed","link":"https://in.indeed.com/viewjob?jk=44b725b689aa3123"},
        {"title":"Wealth Management Trainee","company":"Mint Global LLC","firm_type":"Wealth / Investment Firm","domain":"Wealth Management","location":"Delhi / Gurugram / Noida","stipend":"Rs.15,000/mo + Incentives","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Internshala","link":"https://internshala.com/internships/finance-internship-in-delhi/"},
        {"title":"Financial Audit Intern","company":"PwC India","firm_type":"Big 4 / MNC","domain":"Audit / Assurance","location":"Delhi NCR","stipend":"As per norms","duration":"Structured","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Jobaaj","link":"https://www.jobaaj.com/job/pricewaterhousecoopers-pwc-financial-audit-intern-delhi-ncr-0-to-1-years-819084"},
        {"title":"Investment Banking Intern","company":"HSBC India","firm_type":"MNC Investment Bank","domain":"Investment Banking","location":"Delhi NCR","stipend":"As per norms","duration":"Structured","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Jobaaj","link":"https://www.jobaaj.com/job/hsbc-investment-banking-intern-delhi-ncr-0-to-1-years-670612"},
        {"title":"Corporate Finance Intern","company":"Barclays","firm_type":"MNC Investment Bank","domain":"Investment Banking","location":"Delhi NCR","stipend":"As per norms","duration":"Structured","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Jobaaj","link":"https://www.jobaaj.com/job/barclays-corporate-finance-intern-delhi-ncr-0-to-1-years-760847"},
        {"title":"Equity Research Intern","company":"Trade Brains","firm_type":"Fintech / Research","domain":"Equity Research","location":"Delhi NCR / Remote","stipend":"Rs.10,000-15,000/mo","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Internshala","link":"https://internshala.com/internships/finance-internship/"},
        {"title":"Founder's Office Intern","company":"Undisclosed D2C Startup","firm_type":"Early-Stage Startup","domain":"Founder's Office","location":"Delhi / Gurugram","stipend":"Rs.10,000-20,000/mo","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Wellfound","link":"https://wellfound.com/role/l/founder-intern/india"},
        {"title":"Chief of Staff Intern","company":"Undisclosed SaaS Startup","firm_type":"Tech Startup","domain":"Founder's Office","location":"Gurugram, Haryana","stipend":"Rs.15,000-25,000/mo","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Wellfound","link":"https://wellfound.com/role/l/chief-of-staff/india"},
        {"title":"Management Trainee – Finance & Strategy","company":"Undisclosed Conglomerate","firm_type":"Large Corporate","domain":"Management Trainee","location":"Delhi NCR","stipend":"As per norms","duration":"12 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/management-trainee-jobs-in-delhi-ncr"},
        {"title":"Strategy & Consulting Intern","company":"Undisclosed Boutique","firm_type":"Management Consulting","domain":"Strategy & Consulting","location":"Delhi / Gurugram","stipend":"Rs.15,000-25,000/mo","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Internshala","link":"https://internshala.com/internships/consulting-internship-in-delhi/"},
        {"title":"Operations & Growth Intern","company":"Undisclosed Startup","firm_type":"Growth-Stage Startup","domain":"Operations / Growth","location":"Delhi / Noida","stipend":"Rs.10,000-15,000/mo","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Wellfound","link":"https://wellfound.com/role/l/operations-intern/india"},
        {"title":"Venture Capital Analyst Intern","company":"Undisclosed VC Firm","firm_type":"Venture Capital","domain":"Venture Capital","location":"Delhi / Gurugram","stipend":"Not disclosed","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Wellfound","link":"https://wellfound.com/role/l/venture-capital-intern/india"},
        {"title":"Credit Analyst Intern","company":"Undisclosed NBFC","firm_type":"NBFC / Lending","domain":"Credit / Debt","location":"Gurugram","stipend":"Rs.10,000-15,000/mo","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/credit-analyst-internship-jobs-in-delhi-ncr"},
        {"title":"Valuation Analyst Intern","company":"Undisclosed Boutique","firm_type":"Boutique Advisory","domain":"Valuation","location":"Delhi / Gurugram","stipend":"Rs.10,000-15,000/mo","duration":"3 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Internshala","link":"https://internshala.com/internships/finance-internship-in-delhi/"},
        {"title":"KPO Financial Analyst Intern","company":"Undisclosed KPO","firm_type":"KPO Finance","domain":"KPO Finance","location":"Noida, UP","stipend":"Rs.10,000-12,000/mo","duration":"3-6 Months","posted":"Active","deadline":"Rolling","status":"Not Applied","platform":"Naukri","link":"https://www.naukri.com/kpo-internship-jobs-in-noida"},
    ]


# ═════════════════════════════════════════════
# SOURCE 8: IIM JOBS
# ═════════════════════════════════════════════
def scrape_iimjobs():
    results = []
    searches = [
        ("finance-internship", "Finance Operations"),
        ("investment-banking-internship", "Investment Banking"),
        ("equity-research-internship", "Equity Research"),
        ("strategy-internship", "Strategy & Consulting"),
        ("consulting-internship", "Strategy & Consulting"),
        ("founders-office-internship", "Founder's Office"),
        ("management-trainee", "Management Trainee"),
        ("operations-internship", "Operations / Growth"),
        ("private-equity-internship", "Private Equity"),
        ("audit-internship", "Audit / Assurance"),
    ]
    for slug, fallback in searches:
        try:
            time.sleep(random.uniform(2, 3))
            url = f"https://www.iimjobs.com/j/{slug}-jobs"
            r   = requests.get(url, headers=get_headers(), timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = (soup.select(".job-container") or
                    soup.select(".job_list_item") or
                    soup.select("li.clearfix"))
            for job in jobs[:8]:
                title   = (job.select_one(".job-title") or
                           job.select_one("h2") or job.select_one("a.job-title"))
                company = (job.select_one(".company-name") or
                           job.select_one(".comp-name") or job.select_one("h3"))
                loc     = job.select_one(".loc") or job.select_one(".location")
                link    = job.select_one("a")
                t = title.get_text(strip=True)   if title   else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                # IIMJobs direct link — job URLs contain /j/ 
                direct_link = url
                job_link = (job.select_one("a[href*='/j/']") or
                            job.select_one("a.job-title") or
                            job.select_one("h2 > a") or
                            job.select_one("a[href*='iimjobs']") or
                            link)
                if job_link and job_link.get("href"):
                    h = job_link["href"]
                    direct_link = h if h.startswith("http") else "https://www.iimjobs.com" + h
                href = direct_link
                d = detect_domain(t, c)
                results.append({
                    "title":     t, "company": c,
                    "firm_type": "Corporate / MNC",
                    "domain":    d if d != "Finance Operations" else fallback,
                    "location":  loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend":   "Not disclosed",
                    "duration":  "3-6 Months",
                    "posted":    "< 7 days",
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "IIMJobs",
                    "link":      href
                })
        except Exception as e:
            print(f"IIMJobs error ({slug}): {e}")
    print(f"IIMJobs: {len(results)} results")
    return results


# ═════════════════════════════════════════════
# SOURCE 9: COMPANY CAREER PAGES (Direct)
# ═════════════════════════════════════════════
def scrape_company_careers():
    """
    Scrape top startup & corporate career pages directly.
    These roles often never appear on job boards.
    """
    results = []

    career_pages = [
        # Fintech / Startups
        ("https://careers.cred.club/openings", "CRED", "Fintech / Payments"),
        ("https://zepto.keka.com/careers", "Zepto", "Operations / Growth"),
        ("https://groww.in/blog/careers", "Groww", "Fintech / Payments"),
        ("https://razorpay.com/jobs/", "Razorpay", "Fintech / Payments"),
        ("https://meesho.io/jobs", "Meesho", "Operations / Growth"),
        ("https://www.phonepe.com/en/careers.html", "PhonePe", "Fintech / Payments"),
        # Consulting / Strategy
        ("https://careers.bcg.com/global/en/search-results?keywords=intern&location=India",
         "BCG", "Strategy & Consulting"),
        ("https://www.mckinsey.com/careers/search-jobs?query=intern&location=India",
         "McKinsey", "Strategy & Consulting"),
        # Finance
        ("https://www.hdfcbank.com/content/bbp/repositories/723fb80a-2dde-42a3-9793-7ae1be57c87f/?folderPath=/footer/Careers/",
         "HDFC Bank", "Investment Banking"),
        ("https://jobs.goldmansachs.com/jobs?search=intern&location=india",
         "Goldman Sachs", "Investment Banking"),
    ]

    # Generic selectors to try across different career pages
    title_selectors   = ["h2", "h3", ".job-title", ".position-title",
                         ".opening-title", "[class*=title]", "a[href*=job]"]
    company_selectors = [".company", "h4", "[class*=company]"]
    link_selectors    = ["a[href*=job]", "a[href*=career]",
                         "a[href*=opening]", "a[href*=position]"]

    for url, company_name, fallback_domain in career_pages:
        try:
            time.sleep(random.uniform(2, 3.5))
            r    = requests.get(url, headers=get_headers(), timeout=20)
            soup = BeautifulSoup(r.text, "html.parser")

            # Try each title selector
            jobs_found = []
            for sel in title_selectors:
                items = soup.select(sel)
                if items:
                    jobs_found = items[:6]
                    break

            for item in jobs_found:
                t = item.get_text(strip=True)
                if len(t) < 5 or len(t) > 120:
                    continue
                if not any(kw in t.lower() for kw in [
                    "intern", "trainee", "analyst", "associate",
                    "finance", "strategy", "operations", "founder"
                ]):
                    continue

                # Find link
                link = item.find("a") or item.find_parent("a")
                if link and link.get("href"):
                    href = (url.split("/")[0] + "//" + url.split("/")[2] + link["href"]
                            if link["href"].startswith("/") else link["href"])
                else:
                    href = url

                d = detect_domain(t, company_name)
                results.append({
                    "title":     t,
                    "company":   company_name,
                    "firm_type": "Direct — Company Career Page",
                    "domain":    d if d != "Finance Operations" else fallback_domain,
                    "location":  "Delhi NCR",
                    "stipend":   "Not disclosed",
                    "duration":  "Not mentioned",
                    "posted":    "Active",
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "Company Careers",
                    "link":      href
                })
        except Exception as e:
            print(f"Career page error ({company_name}): {e}")

    print(f"Company Careers: {len(results)} results")
    return results


# ═════════════════════════════════════════════
# SOURCE 10: UNSTOP (formerly Dare2Compete)
# ═════════════════════════════════════════════
def scrape_unstop():
    """
    Unstop is huge for college students — internships,
    competitions, and case challenges all in one place.
    Especially strong for consulting, strategy, and finance roles.
    """
    results = []
    searches = [
        ("finance", "Finance Operations"),
        ("investment-banking", "Investment Banking"),
        ("strategy", "Strategy & Consulting"),
        ("consulting", "Strategy & Consulting"),
        ("operations", "Operations / Growth"),
        ("founder", "Founder's Office"),
        ("equity-research", "Equity Research"),
        ("management-trainee", "Management Trainee"),
        ("fintech", "Fintech / Payments"),
        ("business-development", "Operations / Growth"),
    ]
    for keyword, fallback in searches:
        try:
            time.sleep(random.uniform(2, 3))
            url = f"https://unstop.com/internships?domain={keyword}&location=Delhi"
            r   = requests.get(url, headers=get_headers(), timeout=20)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = (soup.select(".opportunity-card") or
                    soup.select(".card-wrapper") or
                    soup.select("[class*='opportunityCard']") or
                    soup.select("div[class*='card']"))
            for job in jobs[:8]:
                title   = (job.select_one("[class*='title']") or
                           job.select_one("h2") or job.select_one("h3"))
                company = (job.select_one("[class*='company']") or
                           job.select_one("[class*='org']") or
                           job.select_one("h4"))
                loc     = (job.select_one("[class*='location']") or
                           job.select_one("[class*='city']"))
                stipend = (job.select_one("[class*='stipend']") or
                           job.select_one("[class*='salary']"))
                dur     = (job.select_one("[class*='duration']") or
                           job.select_one("[class*='tenure']"))
                # Direct link extraction
                direct_link = url
                job_link = (job.select_one("a[href*='/internship/']") or
                            job.select_one("a[href*='/opportunity/']") or
                            job.select_one("a[href*='unstop.com']") or
                            job.select_one("a"))
                if job_link and job_link.get("href"):
                    h = job_link["href"]
                    direct_link = h if h.startswith("http") else "https://unstop.com" + h

                t = title.get_text(strip=True)   if title   else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                d = detect_domain(t, c)
                results.append({
                    "title":     t, "company": c,
                    "firm_type": "Startup / Corporate",
                    "domain":    d if d != "Finance Operations" else fallback,
                    "location":  loc.get_text(strip=True)     if loc     else "Delhi NCR",
                    "stipend":   stipend.get_text(strip=True) if stipend else "Not disclosed",
                    "duration":  dur.get_text(strip=True)     if dur     else "Not mentioned",
                    "posted":    "< 7 days",
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "Unstop",
                    "link":      direct_link
                })
        except Exception as e:
            print(f"Unstop error ({keyword}): {e}")
    print(f"Unstop: {len(results)} results")
    return results


# ═════════════════════════════════════════════
# LINKEDIN SEARCH URLs (direct filtered links)
# ═════════════════════════════════════════════
def get_linkedin_urls():
    """
    LinkedIn blocks scraping so we generate pre-filtered
    search URLs that open directly to relevant listings.
    One click = filtered LinkedIn search for that domain.
    """
    base = "https://www.linkedin.com/jobs/search/?keywords="
    loc  = "&location=Delhi%20NCR%2C%20India&f_E=1&f_JT=I&sortBy=DD"
    # f_E=1 = Entry level, f_JT=I = Internship, sortBy=DD = Most recent

    linkedin_searches = [
        ("Founder%27s+Office+Intern",        "Founder's Office"),
        ("Management+Trainee+Finance",        "Management Trainee"),
        ("Investment+Banking+Intern",         "Investment Banking"),
        ("Equity+Research+Intern",            "Equity Research"),
        ("Strategy+Consulting+Intern",        "Strategy & Consulting"),
        ("Operations+Growth+Intern",          "Operations / Growth"),
        ("Financial+Analyst+Intern",          "Finance Operations"),
        ("Private+Equity+Intern",             "Private Equity"),
        ("Venture+Capital+Intern",            "Venture Capital"),
        ("Fintech+Intern",                    "Fintech / Payments"),
        ("Valuation+Analyst+Intern",          "Valuation"),
        ("Credit+Analyst+Intern",             "Credit / Debt"),
        ("Audit+Intern",                      "Audit / Assurance"),
        ("Wealth+Management+Intern",          "Wealth Management"),
        ("Chief+of+Staff+Intern",             "Founder's Office"),
    ]

    results = []
    for query, domain in linkedin_searches:
        results.append({
            "title":     "LinkedIn Search — " + domain,
            "company":   "Multiple Companies",
            "firm_type": "Various",
            "domain":    domain,
            "location":  "Delhi NCR",
            "stipend":   "Various",
            "duration":  "Not mentioned",
            "posted":    "Live Search",
            "deadline":  "Click to see",
            "status":    "Not Applied",
            "platform":  "LinkedIn",
            "link":      base + query + loc
        })
    print(f"LinkedIn URLs: {len(results)} search links generated")
    return results


# ═════════════════════════════════════════════
# SOURCE 11: ADZUNA API (free, reliable, India)
# ═════════════════════════════════════════════
def scrape_adzuna():
    """
    Adzuna Jobs API — free tier, 100 calls/day.
    Returns real job listings aggregated from 50+ sources.
    Much more reliable than scraping as it uses official API.
    """
    results = []
    app_id  = os.environ.get("ADZUNA_APP_ID",  "d69dc243")
    app_key = os.environ.get("ADZUNA_APP_KEY", "d0727414dbf421e1c5b42f97019bb50d")

    searches = [
        ("finance intern",              "Finance Operations"),
        ("investment banking intern",   "Investment Banking"),
        ("equity research intern",      "Equity Research"),
        ("strategy intern",             "Strategy & Consulting"),
        ("consulting intern",           "Strategy & Consulting"),
        ("founder office intern",       "Founder's Office"),
        ("management trainee",          "Management Trainee"),
        ("operations intern",           "Operations / Growth"),
        ("fintech intern",              "Fintech / Payments"),
        ("private equity intern",       "Private Equity"),
        ("venture capital intern",      "Venture Capital"),
        ("audit intern",                "Audit / Assurance"),
        ("valuation intern",            "Valuation"),
        ("credit analyst intern",       "Credit / Debt"),
        ("wealth management intern",    "Wealth Management"),
        ("financial analyst intern",    "Finance Operations"),
        ("growth intern",               "Operations / Growth"),
        ("chief of staff intern",       "Founder's Office"),
    ]

    for query, fallback in searches:
        try:
            time.sleep(random.uniform(1, 1.5))
            url = "https://api.adzuna.com/v1/api/jobs/in/search/1"
            params = {
                "app_id":         app_id,
                "app_key":        app_key,
                "results_per_page": 10,
                "what":           query,
                "where":          "Delhi",
                "distance":       50,        # 50km radius
                "max_days_old":   7,
                "content-type":   "application/json",
                "sort_by":        "date",
            }
            r    = requests.get(url, params=params, timeout=15)
            data = r.json()
            jobs = data.get("results", [])

            for job in jobs:
                title       = job.get("title", f"{fallback} Intern")
                company     = job.get("company", {}).get("display_name", "N/A")
                location    = job.get("location", {}).get("display_name", "Delhi NCR")
                description = job.get("description", "")
                redirect    = job.get("redirect_url", "")
                salary_min  = job.get("salary_min", 0)
                salary_max  = job.get("salary_max", 0)
                created     = job.get("created", "")[:10] if job.get("created") else "Recent"

                # Build stipend string from salary data
                if salary_min and salary_max:
                    # Adzuna returns annual — convert to monthly
                    mo_min = int(salary_min / 12)
                    mo_max = int(salary_max / 12)
                    stipend = "Rs." + str(mo_min) + "-" + str(mo_max) + "/mo"
                elif salary_min:
                    mo = int(salary_min / 12)
                    stipend = "Rs." + str(mo) + "/mo"
                else:
                    stipend = "Not disclosed"

                # Skip if clearly not an internship
                combined = (title + " " + description).lower()
                if not any(x in combined for x in [
                    "intern", "trainee", "fresher", "graduate", "entry",
                    "junior", "associate", "analyst"
                ]):
                    continue

                d = detect_domain(title, company)
                results.append({
                    "title":     title,
                    "company":   company,
                    "firm_type": "Corporate / MNC",
                    "domain":    d if d != "Finance Operations" else fallback,
                    "location":  location,
                    "stipend":   stipend,
                    "duration":  "Not mentioned",
                    "posted":    created,
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "Adzuna",
                    "link":      redirect
                })
        except Exception as e:
            print(f"Adzuna error ({query}): {e}")

    print(f"Adzuna: {len(results)} results")
    return results


# ═════════════════════════════════════════════
# SOURCE 12: CUTSHORT (startup hiring platform)
# ═════════════════════════════════════════════
def scrape_cutshort():
    """
    Cutshort is heavily used by Indian startups for hiring.
    Great source for Founder's Office, Strategy, and Growth roles
    that don't appear on traditional job boards.
    """
    results = []
    searches = [
        ("finance", "Finance Operations"),
        ("investment-banking", "Investment Banking"),
        ("strategy", "Strategy & Consulting"),
        ("founder-office", "Founder's Office"),
        ("operations", "Operations / Growth"),
        ("growth", "Operations / Growth"),
        ("fintech", "Fintech / Payments"),
        ("equity-research", "Equity Research"),
        ("consulting", "Strategy & Consulting"),
        ("management-trainee", "Management Trainee"),
    ]
    for keyword, fallback in searches:
        try:
            time.sleep(random.uniform(2, 3))
            url = f"https://cutshort.io/jobs?type=internship&q={keyword}&location=Delhi"
            r   = requests.get(url, headers=get_headers(), timeout=20)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = (soup.select(".job-card") or
                    soup.select("[class*='JobCard']") or
                    soup.select("[class*='job-listing']") or
                    soup.select("div[data-cy='job-card']"))
            for job in jobs[:8]:
                title   = (job.select_one("[class*='title']") or
                           job.select_one("h2") or job.select_one("h3"))
                company = (job.select_one("[class*='company']") or
                           job.select_one("[class*='org']") or
                           job.select_one("h4"))
                loc     = job.select_one("[class*='location']")
                stipend = (job.select_one("[class*='salary']") or
                           job.select_one("[class*='stipend']"))

                # Direct link
                direct_link = url
                job_link = (job.select_one("a[href*='/jobs/']") or
                            job.select_one("a[href*='/internship/']") or
                            job.select_one("a"))
                if job_link and job_link.get("href"):
                    h = job_link["href"]
                    direct_link = h if h.startswith("http") else "https://cutshort.io" + h

                t = title.get_text(strip=True)   if title   else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                d = detect_domain(t, c)
                results.append({
                    "title":     t, "company": c,
                    "firm_type": "Startup",
                    "domain":    d if d != "Finance Operations" else fallback,
                    "location":  loc.get_text(strip=True)     if loc     else "Delhi NCR",
                    "stipend":   stipend.get_text(strip=True) if stipend else "Not disclosed",
                    "duration":  "Not mentioned",
                    "posted":    "< 7 days",
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "Cutshort",
                    "link":      direct_link
                })
        except Exception as e:
            print(f"Cutshort error ({keyword}): {e}")
    print(f"Cutshort: {len(results)} results")
    return results


# ═════════════════════════════════════════════
# SOURCE 13: HIRECT (direct founder hiring)
# ═════════════════════════════════════════════
def scrape_hirect():
    """
    Hirect connects candidates directly with founders/CEOs.
    Best source for Founder's Office and early-stage startup roles.
    """
    results = []
    searches = [
        ("finance-intern", "Finance Operations"),
        ("strategy-intern", "Strategy & Consulting"),
        ("founder-office", "Founder's Office"),
        ("operations-intern", "Operations / Growth"),
        ("business-development", "Operations / Growth"),
        ("management-trainee", "Management Trainee"),
        ("fintech-intern", "Fintech / Payments"),
        ("growth-intern", "Operations / Growth"),
    ]
    for keyword, fallback in searches:
        try:
            time.sleep(random.uniform(2, 3))
            url = f"https://hirect.in/job-search/{keyword}?city=Delhi"
            r   = requests.get(url, headers=get_headers(), timeout=20)
            soup = BeautifulSoup(r.text, "html.parser")
            jobs = (soup.select(".job-card") or
                    soup.select("[class*='JobCard']") or
                    soup.select("[class*='card']") or
                    soup.select("div[class*='job']"))
            for job in jobs[:8]:
                title   = (job.select_one("[class*='title']") or
                           job.select_one("h2") or job.select_one("h3"))
                company = (job.select_one("[class*='company']") or
                           job.select_one("h4"))
                loc     = job.select_one("[class*='location']")
                stipend = job.select_one("[class*='salary']")

                direct_link = url
                job_link = (job.select_one("a[href*='/job/']") or
                            job.select_one("a[href*='/jobs/']") or
                            job.select_one("a"))
                if job_link and job_link.get("href"):
                    h = job_link["href"]
                    direct_link = h if h.startswith("http") else "https://hirect.in" + h

                t = title.get_text(strip=True)   if title   else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"
                d = detect_domain(t, c)
                results.append({
                    "title":     t, "company": c,
                    "firm_type": "Startup (Direct Founder)",
                    "domain":    d if d != "Finance Operations" else fallback,
                    "location":  loc.get_text(strip=True)     if loc     else "Delhi NCR",
                    "stipend":   stipend.get_text(strip=True) if stipend else "Not disclosed",
                    "duration":  "Not mentioned",
                    "posted":    "< 7 days",
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "Hirect",
                    "link":      direct_link
                })
        except Exception as e:
            print(f"Hirect error ({keyword}): {e}")
    print(f"Hirect: {len(results)} results")
    return results


# ═════════════════════════════════════════════
# SOURCE 14: GOOGLE JOBS (aggregates 50+ sites)
# ═════════════════════════════════════════════
def scrape_google_jobs():
    """
    Google Jobs aggregates listings from ALL major job boards.
    One search = results from Naukri, Indeed, Internshala, etc.
    Uses Google's htl jobs endpoint which returns structured data.
    """
    results = []
    searches = [
        ("finance internship Delhi NCR",           "Finance Operations"),
        ("investment banking internship Delhi",     "Investment Banking"),
        ("equity research internship Delhi",        "Equity Research"),
        ("founder office internship Delhi",         "Founder's Office"),
        ("strategy consulting internship Delhi",    "Strategy & Consulting"),
        ("management trainee finance Delhi",        "Management Trainee"),
        ("operations growth internship Delhi",      "Operations / Growth"),
        ("private equity internship Delhi",         "Private Equity"),
        ("fintech internship Delhi NCR",            "Fintech / Payments"),
        ("venture capital internship Delhi",        "Venture Capital"),
        ("valuation analyst internship Delhi",      "Valuation"),
        ("audit internship Delhi NCR",              "Audit / Assurance"),
        ("credit analyst internship Delhi",         "Credit / Debt"),
        ("wealth management internship Delhi",      "Wealth Management"),
        ("chief of staff internship Delhi startup", "Founder's Office"),
    ]

    # Google Jobs uses a special header for structured job results
    google_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-IN,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    for query, fallback in searches:
        try:
            time.sleep(random.uniform(3, 5))  # Google needs longer delays
            encoded = requests.utils.quote(query)
            url = f"https://www.google.com/search?q={encoded}&ibp=htl;jobs&hl=en-IN&gl=in"
            r   = requests.get(url, headers=google_headers, timeout=20)
            soup = BeautifulSoup(r.text, "html.parser")

            # Google Jobs structured data selectors
            jobs = (soup.select("[data-ved][class*='iFjolb']") or
                    soup.select(".PwjeAc") or
                    soup.select("[jscontroller*='job']") or
                    soup.select(".g") or
                    soup.select("[class*='job']"))

            for job in jobs[:6]:
                title   = (job.select_one("[class*='BjJfJf']") or
                           job.select_one("[class*='title']") or
                           job.select_one("h3") or job.select_one("h2"))
                company = (job.select_one("[class*='vNEEBe']") or
                           job.select_one("[class*='company']") or
                           job.select_one("[class*='subtitle']"))
                loc     = (job.select_one("[class*='Qk80Jf']") or
                           job.select_one("[class*='location']"))
                link    = job.select_one("a[href]")

                t = title.get_text(strip=True)   if title   else f"{fallback} Intern"
                c = company.get_text(strip=True) if company else "N/A"

                # Skip generic/irrelevant results
                if len(t) < 5 or len(t) > 150:
                    continue
                if not any(x in t.lower() for x in [
                    "intern", "trainee", "analyst", "associate",
                    "finance", "strategy", "founder", "officer"
                ]):
                    continue

                # Build direct link
                direct_link = ""
                if link and link.get("href"):
                    h = link["href"]
                    if h.startswith("/url?q="):
                        # Google redirect — extract actual URL
                        import urllib.parse
                        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(h).query)
                        direct_link = parsed.get("q", [h])[0]
                    elif h.startswith("http"):
                        direct_link = h
                    else:
                        direct_link = "https://www.google.com" + h

                if not direct_link:
                    direct_link = "https://www.google.com/search?q=" + requests.utils.quote(t + " " + c + " internship apply")

                d = detect_domain(t, c)
                results.append({
                    "title":     t, "company": c,
                    "firm_type": "Various (Google Jobs)",
                    "domain":    d if d != "Finance Operations" else fallback,
                    "location":  loc.get_text(strip=True) if loc else "Delhi NCR",
                    "stipend":   "Not disclosed",
                    "duration":  "Not mentioned",
                    "posted":    "< 7 days",
                    "deadline":  "Not mentioned",
                    "status":    "Not Applied",
                    "platform":  "Google Jobs",
                    "link":      direct_link
                })
        except Exception as e:
            print(f"Google Jobs error ({query}): {e}")

    print(f"Google Jobs: {len(results)} results")
    return results

# ═════════════════════════════════════════════
# BUILD EXCEL  (now with Fit Score column)
# ═════════════════════════════════════════════
def build_excel(internships, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Internships"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font   = Font(name="Arial", size=9)
    link_font   = Font(name="Arial", size=9, color="1558BB", underline="single")
    thin = Border(
        left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),  bottom=Side(style="thin", color="BFBFBF")
    )
    today = date.today().strftime("%B %d, %Y")

    # Title row
    ws.merge_cells("A1:P1")
    ws["A1"] = f"Finance & Management Internships — Delhi NCR | {today} | Sources: Internshala + Indeed + Glassdoor + Naukri + Foundit + Jobaaj + Wellfound"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", start_color="D6E4F0")
    ws.row_dimensions[1].height = 30

    # Legend row
    ws.merge_cells("A2:P2")
    ws["A2"] = ("Color legend — Red: IB | Orange: Equity/Valuation | Purple: PE/VC/HF | "
                "Green: Wealth/AM/FP&A | Blue: Fintech/Credit/KPO | Brown: Audit/Compliance | "
                "Crimson: Founder's Office | Gold: Mgmt Trainee | Steel Blue: Strategy | Sage: Ops/Growth  "
                "★ AI Fit Score: 9-10 = Perfect | 7-8 = Strong | 5-6 = Moderate | 1-4 = Weak")
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="7F6000")
    ws["A2"].fill = PatternFill("solid", start_color="FFF2CC")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = ["#", "Role / Internship Title", "Company", "Tier", "Firm Type", "Domain",
               "Location", "Stipend", "Duration", "Posted", "Deadline", "Platform",
               "Apply Link", "Status", "AI Fit Score", "Why It Fits (AI)", "Notes"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[3].height = 30

    # Sort by fit_score descending, then domain
    internships_sorted = sorted(internships, key=lambda x: (-x.get("fit_score", 5), x.get("domain", "Z")))

    for i, job in enumerate(internships_sorted, 1):
        r       = i + 3
        domain  = job.get("domain", "Finance Operations")
        color   = DOMAIN_COLORS.get(domain, "404040")
        row_bg  = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        score   = job.get("fit_score", 5)

        # Score badge color
        if score >= 9:   score_bg = "C6EFCE"; score_fc = "276221"   # green
        elif score >= 7: score_bg = "FFEB9C"; score_fc = "9C5700"   # yellow
        elif score >= 5: score_bg = "DDEBF7"; score_fc = "1F4E79"   # blue
        else:            score_bg = "FFE0E0"; score_fc = "9C0006"   # red

        values = [i, job["title"], job["company"], job["firm_type"], job["domain"],
                  job["location"], job["stipend"], job["duration"], job["posted"],
                  job["deadline"], job["platform"], job["link"],
                  job["status"], f"{score}/10", job.get("fit_reason", ""), ""]

        for col, val in enumerate(values, 1):
            cell = ws.cell(r, col, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            if col == 2:    # Title
                cell.font = Font(name="Arial", size=9, bold=True, color=color)
                cell.fill = PatternFill("solid", start_color=row_bg)
            elif col == 5:  # Domain
                cell.font = Font(name="Arial", size=9, color=color)
                cell.fill = PatternFill("solid", start_color=row_bg)
            elif col == 12: # Link
                cell.font = link_font
                cell.fill = PatternFill("solid", start_color=row_bg)
                cell.hyperlink = str(val)
            elif col == 13: # Status
                cell.font = Font(name="Arial", size=9, bold=True, color="375623")
                cell.fill = PatternFill("solid", start_color="E2EFDA")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 14: # AI Fit Score
                cell.font = Font(name="Arial", size=10, bold=True, color=score_fc)
                cell.fill = PatternFill("solid", start_color=score_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 15: # Why It Fits
                cell.font = Font(name="Arial", size=8, italic=True, color="404040")
                cell.fill = PatternFill("solid", start_color=row_bg)
            else:
                cell.font = data_font
                cell.fill = PatternFill("solid", start_color=row_bg)
        ws.row_dimensions[r].height = 40

    col_widths = [4, 38, 26, 10, 24, 24, 22, 20, 14, 12, 18, 14, 50, 16, 12, 38, 20]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:P{3 + len(internships)}"

    # ── Sheet 2: Top Picks (fit score >= 7) ──────────────────────────
    ws_top = wb.create_sheet("🌟 Top Picks")
    ws_top.merge_cells("A1:P1")
    ws_top["A1"] = f"Top Picks for Tanishq — AI Fit Score ≥ 7 — {today}"
    ws_top["A1"].font = Font(name="Arial", bold=True, size=12, color="276221")
    ws_top["A1"].fill = PatternFill("solid", start_color="C6EFCE")
    ws_top["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_top.row_dimensions[1].height = 28

    for col, h in enumerate(headers, 1):
        cell = ws_top.cell(2, col, h)
        cell.font = header_font
        cell.fill = PatternFill("solid", start_color="276221")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws_top.row_dimensions[2].height = 28

    top_jobs = [j for j in internships_sorted if j.get("fit_score", 0) >= 7]
    for i, job in enumerate(top_jobs, 1):
        r      = i + 2
        domain = job.get("domain", "Finance Operations")
        color  = DOMAIN_COLORS.get(domain, "404040")
        score  = job.get("fit_score", 7)
        row_bg = "F0FFF0" if i % 2 == 0 else "FFFFFF"
        if score >= 9:   score_bg = "C6EFCE"; score_fc = "276221"
        elif score >= 7: score_bg = "FFEB9C"; score_fc = "9C5700"
        else:            score_bg = "DDEBF7"; score_fc = "1F4E79"

        values = [i, job["title"], job["company"], job["firm_type"], job["domain"],
                  job["location"], job["stipend"], job["duration"], job["posted"],
                  job["deadline"], job["platform"], job["link"],
                  job["status"], f"{score}/10", job.get("fit_reason", ""), ""]
        for col, val in enumerate(values, 1):
            cell = ws_top.cell(r, col, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col == 2:
                cell.font = Font(name="Arial", size=9, bold=True, color=color)
                cell.fill = PatternFill("solid", start_color=row_bg)
            elif col == 12:
                cell.font = link_font
                cell.fill = PatternFill("solid", start_color=row_bg)
                cell.hyperlink = str(val)
            elif col == 14:
                cell.font = Font(name="Arial", size=10, bold=True, color=score_fc)
                cell.fill = PatternFill("solid", start_color=score_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 15:
                cell.font = Font(name="Arial", size=8, italic=True, color="404040")
                cell.fill = PatternFill("solid", start_color=row_bg)
            else:
                cell.font = data_font
                cell.fill = PatternFill("solid", start_color=row_bg)
        ws_top.row_dimensions[r].height = 40

    for col, w in enumerate(col_widths, 1):
        ws_top.column_dimensions[get_column_letter(col)].width = w
    ws_top.freeze_panes = "A3"


    # ── Sheet 3: 3-Month Internships ─────────────────────────────────
    def is_3_month(d_str):
        import re
        d = d_str.lower().strip()
        if any(x in d for x in ["not mentioned", "not disclosed", "n/a", "flexible"]):
            return False
        nums = list(map(int, re.findall(r"\d+", d)))
        if not nums:
            return False
        if "month" in d:
            return min(nums) <= 3 and max(nums) <= 3
        if "week" in d:
            return min(nums) <= 12 and max(nums) <= 12
        return False

    def is_4_month(d_str):
        import re
        d = d_str.lower().strip()
        if any(x in d for x in ["not mentioned", "not disclosed", "n/a", "flexible"]):
            return False
        nums = list(map(int, re.findall(r"\d+", d)))
        if not nums:
            return False
        if "month" in d:
            return 4 in nums or max(nums) == 4
        if "week" in d:
            return min(nums) >= 13 and max(nums) <= 16
        return False

    for sheet_label, sheet_color, header_color, filter_fn, sheet_title in [
        ("3-Month Internships", "FFF2CC", "BF8F00",
         is_3_month, f"3-Month Internships — {today}"),
        ("4-Month Internships", "DDEBF7", "1D6A96",
         is_4_month, f"4-Month Internships — {today}"),
    ]:
        ws_dur = wb.create_sheet(sheet_label)
        ws_dur.merge_cells("A1:P1")
        ws_dur["A1"] = sheet_title
        ws_dur["A1"].font = Font(name="Arial", bold=True, size=12, color=header_color)
        ws_dur["A1"].fill = PatternFill("solid", start_color=sheet_color)
        ws_dur["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dur.row_dimensions[1].height = 28

        for col, h in enumerate(headers, 1):
            cell = ws_dur.cell(2, col, h)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", start_color=header_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        ws_dur.row_dimensions[2].height = 28

        filtered = [j for j in internships_sorted if filter_fn(j.get("duration", ""))]

        if not filtered:
            ws_dur.merge_cells("A3:P3")
            ws_dur["A3"] = "No listings found for this duration in today's run."
            ws_dur["A3"].font = Font(name="Arial", italic=True, size=10, color="7F7F7F")
            ws_dur["A3"].alignment = Alignment(horizontal="center", vertical="center")
            ws_dur.row_dimensions[3].height = 30
        else:
            for i, job in enumerate(filtered, 1):
                r      = i + 2
                domain = job.get("domain", "Finance Operations")
                color  = DOMAIN_COLORS.get(domain, "404040")
                row_bg = "FFFEF0" if sheet_label == "3-Month Internships" else "F0F6FF"
                row_bg = row_bg if i % 2 != 0 else "FFFFFF"
                score  = job.get("fit_score", 5)
                if score >= 9:   score_bg = "C6EFCE"; score_fc = "276221"
                elif score >= 7: score_bg = "FFEB9C"; score_fc = "9C5700"
                elif score >= 5: score_bg = "DDEBF7"; score_fc = "1F4E79"
                else:            score_bg = "FFE0E0"; score_fc = "9C0006"

                values = [i, job["title"], job["company"], job["firm_type"], job["domain"],
                          job["location"], job["stipend"], job["duration"], job["posted"],
                          job["deadline"], job["platform"], job["link"],
                          job["status"], f"{score}/10", job.get("fit_reason", ""), ""]

                for col, val in enumerate(values, 1):
                    cell = ws_dur.cell(r, col, val)
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if col == 2:
                        cell.font = Font(name="Arial", size=9, bold=True, color=color)
                        cell.fill = PatternFill("solid", start_color=row_bg)
                    elif col == 12:
                        cell.font = Font(name="Arial", size=9, color="1558BB", underline="single")
                        cell.fill = PatternFill("solid", start_color=row_bg)
                        cell.hyperlink = str(val)
                    elif col == 13:
                        cell.font = Font(name="Arial", size=9, bold=True, color="375623")
                        cell.fill = PatternFill("solid", start_color="E2EFDA")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col == 14:
                        cell.font = Font(name="Arial", size=10, bold=True, color=score_fc)
                        cell.fill = PatternFill("solid", start_color=score_bg)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col == 15:
                        cell.font = Font(name="Arial", size=8, italic=True, color="404040")
                        cell.fill = PatternFill("solid", start_color=row_bg)
                    else:
                        cell.font = Font(name="Arial", size=9)
                        cell.fill = PatternFill("solid", start_color=row_bg)
                ws_dur.row_dimensions[r].height = 40

        for col, w in enumerate(col_widths, 1):
            ws_dur.column_dimensions[get_column_letter(col)].width = w
        ws_dur.freeze_panes = "A3"
        ws_dur.auto_filter.ref = f"A2:P{2 + max(len(filtered), 1)}"

    # ── Sheet 5: Domain Summary ───────────────────────────────────────
    ws2 = wb.create_sheet("Domain Summary")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = f"Internship Count by Domain — {today}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws2["A1"].fill = PatternFill("solid", start_color="D6E4F0")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28
    for col, h in enumerate(["Domain", "Count", "Source Platforms"], 1):
        cell = ws2.cell(2, col, h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = thin

    domain_counts, domain_platforms = {}, {}
    for job in internships:
        d = job.get("domain", "Finance Operations")
        p = job.get("platform", "")
        domain_counts[d] = domain_counts.get(d, 0) + 1
        domain_platforms.setdefault(d, set()).add(p)

    for row_idx, (domain, count) in enumerate(sorted(domain_counts.items(), key=lambda x: -x[1]), 3):
        color = DOMAIN_COLORS.get(domain, "404040")
        bg = "F5F5F5" if row_idx % 2 == 0 else "FFFFFF"
        c1 = ws2.cell(row_idx, 1, domain);  c1.font = Font(name="Arial", size=10, bold=True, color=color)
        c2 = ws2.cell(row_idx, 2, count);   c2.font = Font(name="Arial", size=10, bold=True); c2.alignment = Alignment(horizontal="center")
        c3 = ws2.cell(row_idx, 3, ", ".join(sorted(domain_platforms.get(domain, [])))); c3.font = Font(name="Arial", size=9)
        for c in [c1, c2, c3]:
            c.fill = PatternFill("solid", start_color=bg); c.border = thin

    last = len(domain_counts) + 3
    for col, val in enumerate(["TOTAL", f"=SUM(B3:B{last-1})", ""], 1):
        cell = ws2.cell(last, col, val)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill("solid", start_color="D6E4F0"); cell.border = thin
        if col == 2: cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 40

    # ── Sheet 6: Application Tracker ─────────────────────────────────
    ws3 = wb.create_sheet("Application Tracker")
    ws3.merge_cells("A1:H1")
    ws3["A1"] = f"My Application Tracker — {today}"
    ws3["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws3["A1"].fill = PatternFill("solid", start_color="D6E4F0")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28
    for col, h in enumerate(["#", "Company", "Role", "Domain", "Applied Date", "Status", "Interview Date", "Notes"], 1):
        cell = ws3.cell(2, col, h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin
    ws3.row_dimensions[2].height = 28
    status_map = [
        ("Not Applied","F2F2F2"),("Applied","BDD7EE"),("In Review","FFF2CC"),
        ("Interview Scheduled","E2EFDA"),("Offer Received","C6EFCE"),("Rejected","FFE0E0"),
    ]
    for i in range(3, 53):
        row_bg = "FFFFFF" if i % 2 == 0 else "F9F9F9"
        for col in range(1, 9):
            cell = ws3.cell(i, col)
            cell.fill = PatternFill("solid", start_color=row_bg)
            cell.border = thin; cell.font = data_font
            if col == 1: cell.value = i - 2; cell.alignment = Alignment(horizontal="center")
    ws3.merge_cells("A55:H55")
    ws3["A55"] = "STATUS LEGEND"
    ws3["A55"].font = Font(name="Arial", bold=True, color="FFFFFF")
    ws3["A55"].fill = PatternFill("solid", start_color="1F4E79")
    ws3["A55"].alignment = Alignment(horizontal="center")
    for idx, (status, color) in enumerate(status_map, 56):
        cell = ws3.cell(idx, 1, status)
        cell.font = Font(name="Arial", bold=True, size=9)
        cell.fill = PatternFill("solid", start_color=color); cell.border = thin
        cell.alignment = Alignment(horizontal="center")
        ws3.merge_cells(f"B{idx}:H{idx}")
        ws3.cell(idx, 2).fill = PatternFill("solid", start_color=color)
        ws3.cell(idx, 2).border = thin
    for col, w in zip(range(1, 9), [4, 28, 36, 24, 16, 22, 18, 36]):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A3"

    # ── Sheet 7: LinkedIn Search URLs ────────────────────────────────
    ws_li = wb.create_sheet("LinkedIn Search URLs")
    ws_li.merge_cells("A1:D1")
    ws_li["A1"] = f"LinkedIn Direct Search Links — {today} (Click to open filtered results)"
    ws_li["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws_li["A1"].fill = PatternFill("solid", start_color="0A66C2")
    ws_li["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_li.row_dimensions[1].height = 28

    li_headers = ["Domain", "Search Link", "Filter Applied", "Tip"]
    li_tips = {
        "Founder's Office": "Sort by Most Recent, apply within 24hrs",
        "Investment Banking": "Target boutique IB firms for better response rate",
        "Equity Research": "Mention SEBI NISM cert in first line of message",
        "Strategy & Consulting": "Reference McKinsey Forward in your message",
        "Operations / Growth": "Highlight 23 Ventures hackathon experience",
        "Management Trainee": "Apply to FMCG and conglomerate MT programs",
        "Private Equity": "Cold email is better than applying online for PE",
        "Venture Capital": "Mention startup credits from 23 Ventures",
        "Fintech / Payments": "Focus on product-facing finance roles",
        "Valuation": "Highlight DCF and financial modeling in headline",
        "Credit / Debt": "NBFC roles are more accessible than bank roles",
        "Audit / Assurance": "Big 4 has structured internship programs",
        "Wealth Management": "SEBI NISM cert is a strong differentiator",
        "Finance Operations": "Good stepping stone — builds domain experience",
    }

    for col, h in enumerate(li_headers, 1):
        cell = ws_li.cell(2, col, h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="0A66C2")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws_li.row_dimensions[2].height = 25

    li_jobs = [j for j in internships if j.get("platform") == "LinkedIn"]
    for i, job in enumerate(li_jobs, 1):
        r      = i + 2
        row_bg = "EEF3FB" if i % 2 != 0 else "FFFFFF"
        domain = job.get("domain", "")
        tip    = li_tips.get(domain, "Apply within 24hrs of posting for best response")

        vals = [domain, job.get("link", ""), "Entry Level | Internship | Delhi NCR | Recent", tip]
        for col, val in enumerate(vals, 1):
            cell = ws_li.cell(r, col, val)
            cell.font = (Font(name="Arial", size=9, color="0A66C2", underline="single")
                        if col == 2 else Font(name="Arial", size=9))
            cell.fill = PatternFill("solid", start_color=row_bg)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col == 2 and val:
                cell.hyperlink = val
        ws_li.row_dimensions[r].height = 30

    ws_li.column_dimensions["A"].width = 26
    ws_li.column_dimensions["B"].width = 60
    ws_li.column_dimensions["C"].width = 34
    ws_li.column_dimensions["D"].width = 48
    ws_li.freeze_panes = "A3"

    wb.save(filepath)
    print(f"Excel saved: {filepath}")


# ═════════════════════════════════════════════
# DURATION FILTER — max 4 months
# ═════════════════════════════════════════════
def filter_by_duration(jobs):
    import re

    def is_acceptable(d_str):
        d = d_str.lower().strip()
        # Keep if duration not specified — benefit of doubt
        if any(x in d for x in ["not mentioned", "not disclosed", "n/a",
                                  "rolling", "flexible", "part-time",
                                  "as per", "structured"]):
            return True
        nums = list(map(int, re.findall(r"\d+", d)))
        if not nums:
            return True
        min_num = min(nums)
        if "year" in d:
            return False          # 1 year+ always rejected
        if "month" in d:
            return min_num <= 4   # 3-6 months → min=3 → keep; 6-12 → min=6 → reject
        if "week" in d:
            return min_num <= 16  # 16 weeks = 4 months
        return min_num <= 4

    kept    = [j for j in jobs if is_acceptable(j.get("duration", "N/A"))]
    removed = len(jobs) - len(kept)
    print(f"Duration filter: removed {removed} listings over 4 months, {len(kept)} kept.")
    return kept

# ═════════════════════════════════════════════
# SEND EMAIL
# ═════════════════════════════════════════════
def send_email(filepath, internships, new_count):
    today_str = date.today().strftime("%B %d, %Y")
    total     = len(internships)
    top_picks = [j for j in internships if j.get("fit_score", 0) >= 9]

    domain_counts   = {}
    platform_counts = {}
    for job in internships:
        d = job.get("domain", "Finance Operations")
        p = job.get("platform", "Other")
        domain_counts[d]   = domain_counts.get(d, 0) + 1
        platform_counts[p] = platform_counts.get(p, 0) + 1

    domain_lines   = "\n".join([f"   {d}: {c}" for d, c in sorted(domain_counts.items(), key=lambda x: -x[1])])
    platform_lines = "\n".join([f"   {p}: {c}" for p, c in sorted(platform_counts.items(), key=lambda x: -x[1])])

    # Top picks summary for email body
    newline = "\n"
    top_lines = ""
    for j in sorted(top_picks, key=lambda x: -x.get("fit_score", 0))[:5]:
        score    = j["fit_score"]
        title    = j["title"]
        company  = j["company"]
        location = j["location"]
        reason   = j.get("fit_reason", "")
        link     = j["link"]
        top_lines += "   [" + str(score) + "/10] " + title + " @ " + company + " — " + location + newline
        top_lines += "          -> " + reason + newline
        top_lines += "          -> " + link + newline + newline

    msg = MIMEMultipart()
    msg["From"]    = YOUR_EMAIL
    msg["To"]      = SEND_TO_EMAIL
    msg["Subject"] = f"🎯 {new_count} NEW Internships for Tanishq — {today_str} — {total} Total"

    no_picks_msg = "   No 9+ listings this run — check Sheet 1 for 7+ scores."
    top_section  = top_lines if top_lines else no_picks_msg

    body = f"""Hi Tanishq,

Your Internship Bot v8.0 report is ready!

DATE      : {today_str}
NEW TODAY : {new_count} listings (not seen before)
TOTAL     : {total} internships (including recurring)
LOCATION  : Delhi / Gurugram / Noida / Greater Noida
SCHEDULE  : Runs every day at 8:00 AM IST
SOURCES   : 14 platforms + Adzuna API (100+ daily)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🌟 TOP PICKS FOR YOU (AI Score 9+/10)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{top_section}

SOURCES SCRAPED:
{platform_lines}

DOMAIN BREAKDOWN:
{domain_lines}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXCEL HAS 7 SHEETS:
  Sheet 1 — All Internships (sorted by AI Fit Score, color coded)
  Sheet 2 — Top Picks (Fit Score 7+, best matches for you)
  Sheet 3 — 3-Month Internships (short, focused internships)
  Sheet 4 — 4-Month Internships (slightly longer internships)
  Sheet 5 — Domain Summary (count per domain + platforms)
  Sheet 6 — Application Tracker (track your progress)
  Sheet 7 — LinkedIn Search URLs (direct filtered links)

AI FIT SCORE is based on your resume:
  BBA Finance & Banking | SEBI NISM | McKinsey Forward
  Yhills Finance Intern | 23 Ventures Founder Fellow
  Skills: Excel, Power BI, Financial Modeling, Strategy

HOW TO USE:
  1. Start with Sheet 2 (Top Picks) — highest fit first
  2. Click Apply Link to go directly to the listing
  3. Update STATUS column after applying
  4. Use Sheet 4 to track interview progress

Good luck!
-- Internship Bot v8.0 (GitHub Actions)
   Sources: Internshala + Indeed + Glassdoor + Naukri + Foundit + Jobaaj + Wellfound + IIMJobs + Company Careers + Unstop + LinkedIn URLs + Adzuna API + Cutshort + Hirect + Google Jobs
   Dedup: Google Sheets | AI Scoring: Claude API
   Schedule: Every day at 8:00 AM IST
"""
    msg.attach(MIMEText(body, "plain"))
    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        fname = os.path.basename(filepath)
        part.add_header("Content-Disposition", f"attachment; filename={fname}")
        msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(YOUR_EMAIL, YOUR_PASSWORD)
        server.sendmail(YOUR_EMAIL, SEND_TO_EMAIL, msg.as_string())
    print("Email sent!")


# ═════════════════════════════════════════════
# COMPANY TIER TAGGING
# ═════════════════════════════════════════════
TIER1_COMPANIES = {
    # Big 4 & Global Consulting
    "pwc", "deloitte", "kpmg", "ey", "ernst", "mckinsey", "bcg", "bain",
    "accenture", "oliver wyman", "roland berger",
    # Global Banks & Finance
    "goldman sachs", "morgan stanley", "jp morgan", "jpmorgan", "hsbc",
    "barclays", "citibank", "citi", "deutsche bank", "ubs", "credit suisse",
    "blackrock", "vanguard", "fidelity",
    # Top Indian Finance
    "kotak", "hdfc", "icici", "axis bank", "sbi", "sebi", "rbi",
    "edelweiss", "motilal oswal", "iifl", "angel broking", "zerodha",
    # Top Indian Startups (unicorns)
    "cred", "zepto", "groww", "razorpay", "phonepe", "paytm", "meesho",
    "swiggy", "zomato", "ola", "nykaa", "mamaearth", "boat",
    "dream11", "games24x7", "mpl", "unacademy", "byju",
    # PE / VC
    "sequoia", "accel", "tiger global", "softbank", "general atlantic",
    "warburg pincus", "carlyle", "blackstone", "kkr",
}

TIER2_COMPANIES = {
    # Mid-tier consulting & finance
    "wns", "genpact", "mphasis", "hexaware", "niit", "firstsource",
    "ujjivan", "bandhan", "au small finance", "suryoday",
    "avendus", "o3 capital", "veda corporate", "anand rathi",
    "sharekhan", "5paisa", "upstox", "coin", "dhan",
    # Mid startups
    "travelclan", "cleartax", "zoho", "freshworks", "chargebee",
    "postman", "browserstack", "darwinbox", "leadsquared",
    "lendingkart", "capital float", "indifi", "yubi", "credavenue",
}

def tag_company_tier(company_name):
    c = company_name.lower().strip()
    if any(t in c for t in TIER1_COMPANIES):
        return "Tier 1"
    if any(t in c for t in TIER2_COMPANIES):
        return "Tier 2"
    return "Tier 3"

def apply_tier_tags(jobs):
    for job in jobs:
        job["tier"] = tag_company_tier(job.get("company", ""))
    tier_counts = {}
    for job in jobs:
        t = job["tier"]
        tier_counts[t] = tier_counts.get(t, 0) + 1
    print(f"Tier tagging: {tier_counts}")
    return jobs


# ═════════════════════════════════════════════
# STIPEND FILTER — minimum Rs.8,000/month
# ═════════════════════════════════════════════
def filter_by_stipend(jobs):
    """
    Remove listings with stipend clearly below Rs.8,000/mo.
    Keep if stipend is not disclosed (can't confirm it's low).
    """
    import re

    def is_acceptable_stipend(stipend_str):
        s = stipend_str.lower().strip()

        # Keep if not disclosed — benefit of doubt
        if any(x in s for x in ["not disclosed", "not mentioned", "as per",
                                  "negotiable", "competitive", "as per norms",
                                  "industry standard", "incentive"]):
            return True

        # Extract all numbers
        nums = re.findall(r"[\d,]+", s)
        nums = [int(n.replace(",", "")) for n in nums if n.replace(",", "").isdigit()]
        if not nums:
            return True

        max_num = max(nums)

        # If number looks like monthly stipend (< 1,00,000)
        if max_num < 100000:
            return max_num >= 8000

        # If annual (> 1,00,000) — convert to monthly
        monthly = max_num / 12
        return monthly >= 8000

    kept    = [j for j in jobs if is_acceptable_stipend(j.get("stipend", "Not disclosed"))]
    removed = len(jobs) - len(kept)
    print(f"Stipend filter: removed {removed} listings below Rs.8,000/mo, {len(kept)} kept.")
    return kept


# ═════════════════════════════════════════════
# WEEKLY STRATEGIC REPORT (Sundays only)
# ═════════════════════════════════════════════
def send_weekly_report(internships, gs_client):
    """
    Every Sunday, send a strategic AI-generated report analyzing:
    - Which domains have most opportunities
    - Which platforms performed best
    - Tier 1 company count
    - Strategic recommendations for the coming week
    """
    today_weekday = date.today().weekday()
    if today_weekday != 6:  # 6 = Sunday
        return

    print("Sunday detected — generating weekly strategic report...")
    today_str = date.today().strftime("%B %d, %Y")

    # ── Gather stats ──
    domain_counts   = {}
    platform_counts = {}
    tier_counts     = {}
    high_fit        = []

    for job in internships:
        d = job.get("domain", "Finance Operations")
        p = job.get("platform", "Other")
        t = job.get("tier", "Tier 3")
        domain_counts[d]   = domain_counts.get(d, 0) + 1
        platform_counts[p] = platform_counts.get(p, 0) + 1
        tier_counts[t]     = tier_counts.get(t, 0) + 1
        if job.get("fit_score", 0) >= 8:
            high_fit.append(job)

    top_domains   = sorted(domain_counts.items(),   key=lambda x: -x[1])[:5]
    top_platforms = sorted(platform_counts.items(), key=lambda x: -x[1])[:5]
    tier1_count   = tier_counts.get("Tier 1", 0)
    tier2_count   = tier_counts.get("Tier 2", 0)

    # ── Build stats text for Claude ──
    stats_text = "Domain breakdown: " + ", ".join([d + ": " + str(c) for d, c in top_domains])
    stats_text += ". Platform breakdown: " + ", ".join([p + ": " + str(c) for p, c in top_platforms])
    stats_text += ". Tier 1 companies: " + str(tier1_count)
    stats_text += ". High fit listings (8+): " + str(len(high_fit))
    stats_text += ". Total listings this week: " + str(len(internships))

    top_picks_text = ""
    for j in sorted(high_fit, key=lambda x: -x.get("fit_score", 0))[:10]:
        score   = j.get("fit_score", 0)
        title   = j.get("title", "")
        company = j.get("company", "")
        domain  = j.get("domain", "")
        tier    = j.get("tier", "")
        top_picks_text += str(score) + "/10 — " + title + " @ " + company + " (" + domain + ", " + tier + ")\n"

    # ── Ask Claude for strategic recommendations ──
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    prompt = """You are a career advisor for a college student urgently looking for internships.

CANDIDATE: Tanishq Singhal
- BBA Finance & Banking, 2nd year, IMS UCC Ghaziabad
- SEBI NISM certified, McKinsey Forward Program, IIT Guwahati Winter Consulting Top 10%
- Experience: Finance Intern at Yhills, Founder Fellow at 23 Ventures
- Skills: Financial Modeling, Excel, Power BI, Strategy, Leadership
- Goal: Land internship within 60 days
- Best fit domains: Founder's Office, Strategy & Consulting, Investment Banking, Equity Research

THIS WEEK'S INTERNSHIP DATA:
""" + stats_text + """

TOP LISTINGS THIS WEEK:
""" + top_picks_text + """

Write a concise weekly strategic report (plain text, no markdown) with:
1. WEEK SUMMARY — key observations about this week's listings in 2-3 sentences
2. TOP 3 DOMAINS TO FOCUS ON — which domains to prioritize this week and why
3. BEST PLATFORM THIS WEEK — which platform gave best results and why
4. ACTION PLAN — 5 specific actions Tanishq should take this coming week
5. SKILL GAP ALERT — any skills repeatedly appearing in listings that Tanishq should highlight or develop

Keep it practical, specific, and motivating. Max 400 words."""

    try:
        response = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        report_text = response.content[0].text.strip()
    except Exception as e:
        report_text = "AI report generation failed: " + str(e)
        print(report_text)

    # ── Send separate weekly report email ──
    msg = MIMEMultipart()
    msg["From"]    = YOUR_EMAIL
    msg["To"]      = SEND_TO_EMAIL
    msg["Subject"] = "📊 Weekly Internship Strategy Report — " + today_str

    nl = "\n"
    body = ("Hi Tanishq," + nl + nl +
            "Your Weekly Strategic Internship Report is ready!" + nl + nl +
            "=" * 50 + nl +
            "WEEKLY STATS SNAPSHOT" + nl +
            "=" * 50 + nl +
            "Total listings this week : " + str(len(internships)) + nl +
            "High fit listings (8+/10): " + str(len(high_fit)) + nl +
            "Tier 1 companies         : " + str(tier1_count) + nl +
            "Tier 2 companies         : " + str(tier2_count) + nl + nl +
            "Top domains:" + nl +
            nl.join(["  " + d + ": " + str(c) + " listings" for d, c in top_domains]) + nl + nl +
            "Top platforms:" + nl +
            nl.join(["  " + p + ": " + str(c) + " listings" for p, c in top_platforms]) + nl + nl +
            "=" * 50 + nl +
            "AI STRATEGIC REPORT" + nl +
            "=" * 50 + nl +
            report_text + nl + nl +
            "=" * 50 + nl +
            "TOP LISTINGS TO APPLY THIS WEEK" + nl +
            "=" * 50 + nl +
            top_picks_text + nl +
            "-- Internship Bot v7.0 | Weekly Report every Sunday" + nl)

    msg.attach(MIMEText(body, "plain"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(YOUR_EMAIL, YOUR_PASSWORD)
        server.sendmail(YOUR_EMAIL, SEND_TO_EMAIL, msg.as_string())
    print("Weekly strategic report sent!")

# ═════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════
def run():
    # Schedule guard
    today_weekday = date.today().weekday()
    if today_weekday not in RUN_ON_WEEKDAYS:
        day_name = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"][today_weekday]
        print(f"Skipping — today is {day_name}. Bot runs every day.")
        return

    print("Starting Internship Bot v8.0...")

    # ── Step 1: Scrape all sources ──
    all_jobs = []
    all_jobs += scrape_internshala()
    all_jobs += scrape_indeed()
    all_jobs += scrape_glassdoor()
    all_jobs += scrape_naukri()
    all_jobs += scrape_foundit()
    all_jobs += scrape_jobaaj()
    all_jobs += scrape_wellfound()
    all_jobs += scrape_iimjobs()           # NEW
    all_jobs += scrape_company_careers()
    all_jobs += scrape_unstop()              # NEW
    all_jobs += get_linkedin_urls()
    all_jobs += scrape_adzuna()              # NEW — API based
    all_jobs += scrape_cutshort()            # NEW — startups
    all_jobs += scrape_hirect()              # NEW — founder hiring
    all_jobs += scrape_google_jobs()         # NEW — aggregates 50+ sites
    all_jobs += get_quality_listings()

    # ── Step 2: Local dedup (same run) ──
    seen_local, unique_all = set(), []
    for j in all_jobs:
        key = (j["title"].strip().lower()[:40], j["company"].strip().lower()[:30])
        if key not in seen_local:
            seen_local.add(key)
            unique_all.append(j)
    print(f"After local dedup: {len(unique_all)} unique listings")

    # ── Step 2b: Duration filter (max 4 months) ──
    unique_all = filter_by_duration(unique_all)

    # ── Step 2c: Stipend filter (min Rs.8,000/mo) ──
    unique_all = filter_by_stipend(unique_all)

    # ── Step 2d: Company tier tagging ──
    unique_all = apply_tier_tags(unique_all)

    # ── Step 3: Cross-run dedup via Google Sheets ──
    gs_client              = get_gsheet_client()
    seen_keys, dedup_ws    = load_seen_keys(gs_client)
    new_jobs               = filter_new_only(unique_all, seen_keys)
    new_count              = len(new_jobs)

    # Save new keys back to Google Sheets
    save_new_keys(dedup_ws, new_jobs)

    # For the Excel we show ALL unique (not just new) so context isn't lost,
    # but email subject highlights new count
    jobs_to_report = unique_all

    # ── Step 4: AI Fit Scoring ──
    print("Running AI fit scoring...")
    jobs_to_report = score_jobs_with_ai(jobs_to_report)

    # ── Step 5: Build Excel & Send Email ──
    today_str = date.today().strftime("%Y-%m-%d")
    filepath  = f"/tmp/Internships_Tanishq_{today_str}.xlsx"
    build_excel(jobs_to_report, filepath)
    send_email(filepath, jobs_to_report, new_count)

    # ── Step 6: Weekly strategic report (Sundays only) ──
    send_weekly_report(jobs_to_report, gs_client)

    print(f"Done! {len(jobs_to_report)} total listings, {new_count} new this run.")

if __name__ == "__main__":
    run()
